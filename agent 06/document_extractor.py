"""
Document Extraction Agent
=========================
Receives a list of files (with processing metadata) and a process step definition.
Extracts content from each file using the appropriate tool, tags each section for
relevance to the process step, and returns a structured extraction JSON per file.

Endpoint: POST /extract
Port:     8090
"""

import concurrent.futures
import ast
import json
import logging
import os
import re
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import base64
import io

import anthropic
import google.generativeai as genai
import openai
import PIL.Image
import pdfplumber
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

logger = logging.getLogger("document_extractor")
logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Document Extraction Agent")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

LLM_MODEL              = os.getenv("LLM_MODEL",              "gpt-4o")
TAGGING_MODEL          = os.getenv("TAGGING_MODEL",          "gpt-4o-mini")
VISION_MODEL           = os.getenv("VISION_MODEL",           "gemini-2.0-flash")
ASSET_EXTRACTION_MODEL = os.getenv("ASSET_EXTRACTION_MODEL", "claude-sonnet-4-6")

_openai_client    = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
_anthropic_client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

# Per-request model routing — set in extract_documents, read by _llm_call
_req_context = threading.local()

OUTPUT_DIR = Path("/app/OUTPUT")
OUTPUT_DIR.mkdir(exist_ok=True)

# Sections per batch for tagging — smaller batches are more reliable
TAGGING_BATCH_SIZE = 15

# Retries and backoff for LLM calls
LLM_MAX_RETRIES  = 3
LLM_RETRY_DELAYS = [2, 5, 10]  # seconds between attempts

# Minimum alpha-character ratio to consider a section text meaningful
MIN_ALPHA_RATIO = 0.30

# Maximum sections before falling back to page-level grouping
MAX_SECTIONS_BEFORE_FALLBACK = 80


# ── Section heading patterns ──────────────────────────────────────────────────
# Numbered headings: "1.2 Section Title"
_NUMBERED = re.compile(
    r"^\s*(\d+(?:\.\d+){0,3})\s+([A-Z][A-Za-z0-9 ,\-/&:()\[\]']{2,70})\s*$"
)
# ALL-CAPS headings: must be at least 2 words, ≥8 chars, no repeated single word
# Excludes lines that are pure coordinates, addresses, or label spam
_ALLCAPS = re.compile(r"^\s*([A-Z][A-Z0-9]{1,}(?:\s+[A-Z][A-Z0-9]{1,}){1,})\s*$")

# High-value headings that always force a section boundary regardless of capitalisation.
# These are critical financial/legal sections commonly missed by the ALLCAPS detector.
_FORCED_HEADINGS = re.compile(
    r"^\s*(determination\s+of\s+(funding|supply)|ancillary\s+network\s+services(\s+fees?)?|"
    r"land\s+interests?|method\s+of\s+supply|funding\s+requirements?|"
    r"capital\s+contribution|contestable\s+works|non.contestable\s+works)\s*$",
    re.IGNORECASE,
)


def _is_valid_heading(text: str) -> bool:
    """Extra validation for ALL-CAPS candidate headings."""
    words = text.strip().split()
    if len(words) < 2 or len(text.strip()) < 8:
        return False
    # Reject if more than 60% of words are identical (e.g. "GAS GAS GAS GAS")
    if len(set(words)) / len(words) < 0.5:
        return False
    # Reject if it looks like a street address (contains NSW, VIC, QLD, etc.)
    address_tokens = {"NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT"}
    if address_tokens & set(words):
        return False
    # Reject if it looks like a date (contains month names or years)
    months = {"JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY",
              "AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"}
    if months & set(words):
        return False
    return True


def _alpha_ratio(text: str) -> float:
    """Return fraction of characters that are alphabetic."""
    if not text:
        return 0.0
    alpha = sum(c.isalpha() for c in text)
    return alpha / len(text)


_JSON_FENCE = re.compile(r"```(?:json)?\s*([\s\S]*?)```", re.IGNORECASE)

def _strip_json_comments(text: str) -> str:
    """Remove JS-style // line and /* block */ comments from JSON text, respecting strings."""
    result = []
    i = 0
    in_string = False
    while i < len(text):
        if in_string:
            if text[i] == '\\':
                result.append(text[i])
                i += 1
                if i < len(text):
                    result.append(text[i])
                    i += 1
            elif text[i] == '"':
                result.append(text[i])
                in_string = False
                i += 1
            else:
                result.append(text[i])
                i += 1
        else:
            if text[i] == '"':
                result.append(text[i])
                in_string = True
                i += 1
            elif text[i:i+2] == '//':
                while i < len(text) and text[i] != '\n':
                    i += 1
            elif text[i:i+2] == '/*':
                end = text.find('*/', i + 2)
                i = len(text) if end == -1 else end + 2
            else:
                result.append(text[i])
                i += 1
    return ''.join(result)

def _escape_string_newlines(text: str) -> str:
    """Escape literal newlines/carriage-returns inside JSON double-quoted string values.
    Gemini sometimes embeds raw newlines in string values, making the JSON invalid."""
    result = []
    in_string = False
    i = 0
    while i < len(text):
        c = text[i]
        if in_string:
            if c == '\\' and i + 1 < len(text):
                # Already-escaped sequence — pass through both chars unchanged
                result.append(c)
                result.append(text[i + 1])
                i += 2
                continue
            elif c == '"':
                in_string = False
                result.append(c)
            elif c == '\n':
                result.append('\\n')
            elif c == '\r':
                result.append('\\r')
            else:
                result.append(c)
        else:
            if c == '"':
                in_string = True
            result.append(c)
        i += 1
    return ''.join(result)


def _parse_json_robust(text: str) -> Any:
    """Parse JSON from LLM response, handling fences, JS comments, literal newlines, and single-quoted literals."""
    text = text.strip()
    m = _JSON_FENCE.search(text)
    if m:
        text = m.group(1).strip()
    text = _strip_json_comments(text)

    # Attempt 1: standard JSON parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Attempt 2: repair literal newlines inside string values (common Gemini issue)
    try:
        return json.loads(_escape_string_newlines(text))
    except json.JSONDecodeError:
        pass

    # Attempt 3: Python-style single-quoted dicts (Gemini sometimes returns these)
    py_text = re.sub(r'\bnull\b', 'None', text)
    py_text = re.sub(r'\btrue\b', 'True', py_text)
    py_text = re.sub(r'\bfalse\b', 'False', py_text)
    try:
        return ast.literal_eval(py_text)
    except (ValueError, SyntaxError):
        pass

    # Attempt 4: repair newlines then try Python literal
    try:
        repaired = _escape_string_newlines(py_text)
        return ast.literal_eval(repaired)
    except (ValueError, SyntaxError):
        pass

    raise json.JSONDecodeError("Failed to parse LLM response after all repair attempts", text, 0)


def _llm_call_claude(prompt: str, model: str) -> Any:
    """Call a Claude model via Anthropic API. Returns parsed JSON."""
    last_exc = None
    for attempt in range(LLM_MAX_RETRIES):
        try:
            msg = _anthropic_client.messages.create(
                model=model,
                max_tokens=16000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}],
            )
            return _parse_json_robust(msg.content[0].text)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Claude call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call(prompt: str) -> Any:
    """Call the configured LLM with retry logic. Routes to Claude or OpenAI based on request context."""
    active_model = getattr(_req_context, "model", LLM_MODEL)
    if active_model.startswith("claude-"):
        return _llm_call_claude(prompt, active_model)
    last_exc = None
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = _openai_client.chat.completions.create(
                model=active_model,
                messages=[{"role": "user", "content": prompt}],
                response_format={"type": "json_object"},
                temperature=0,
            )
            return _parse_json_robust(resp.choices[0].message.content)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"LLM call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call_fast(prompt: str) -> Any:
    """Call GPT-4o-mini (fast tagging) with retry logic. Returns parsed JSON."""
    last_exc = None
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = _openai_client.chat.completions.create(
                model=TAGGING_MODEL,
                messages=[{"role": "user", "content": prompt}],
                response_format={"type": "json_object"},
                temperature=0,
            )
            return _parse_json_robust(resp.choices[0].message.content)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Fast LLM call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call_vision(prompt: str, image_bytes: bytes) -> Any:
    """Call Gemini with an inline PNG image + prompt. Returns parsed JSON."""
    img = PIL.Image.open(io.BytesIO(image_bytes))
    last_exc = None
    model = genai.GenerativeModel(VISION_MODEL)
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = model.generate_content([prompt, img])
            return _parse_json_robust(resp.text)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Vision LLM call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call_vision_multi(prompt: str, image_bytes_list: List[bytes]) -> Any:
    """Call Gemini with multiple inline PNG images + prompt. Returns parsed JSON.
    Used for boundary stitching — shows two adjacent chunk images simultaneously."""
    parts: list = [prompt]
    for img_bytes in image_bytes_list:
        parts.append(PIL.Image.open(io.BytesIO(img_bytes)))
    last_exc = None
    model = genai.GenerativeModel(VISION_MODEL)
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = model.generate_content(parts)
            return _parse_json_robust(resp.text)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Multi-image vision call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _vision_page_has_legend(image_bytes_list: List[bytes]) -> bool:
    """Ask the vision model whether a set of page chunk images contain a legend/symbol table.
    Sends chunk images together so the model can assess the page as a whole.
    Returns True if a legend is detected, False on confident NO.
    Falls back to True on error (conservative — avoids silently dropping real entries)."""
    prompt = (
        "You are reviewing image tiles from a large-format engineering drawing. "
        "These tiles show sections of a single page.\n"
        "A LEGEND (also called KEY, SYMBOL TABLE, or NOTATION TABLE) is a structured section "
        "that shows drawn graphical symbols (lines, circles, filled shapes, rectangles) "
        "on the left paired with text labels on the right — "
        "it is NOT a title block, revision table, schedule, or block of notes.\n\n"
        "Do any of these images contain a LEGEND or SYMBOL TABLE with drawn symbols "
        "paired with text labels?\n\n"
        "Reply with a JSON object only: {\"has_legend\": true} or {\"has_legend\": false}"
    )
    parts: list = [prompt]
    for img_bytes in image_bytes_list:
        parts.append(PIL.Image.open(io.BytesIO(img_bytes)))
    model = genai.GenerativeModel(VISION_MODEL)
    try:
        resp = model.generate_content(parts)
        result = _parse_json_robust(resp.text)
        if isinstance(result, dict):
            return bool(result.get("has_legend", True))
        return "true" in str(resp.text).lower()
    except Exception as e:
        logger.warning(f"Page legend presence check failed: {e} — assuming legend present")
        return True


# ── Request / response models ─────────────────────────────────────────────────

class FileEntry(BaseModel):
    filename: str
    filepath: Optional[str] = None
    processing_tool_id: Optional[str] = "tool-extract-pdf-content"
    document_type: Optional[str] = None
    document_category: Optional[str] = None
    content_type: Optional[str] = "text"
    text_quality: Optional[str] = "high"
    page_size: Optional[str] = "A4"
    requires_chunking: Optional[bool] = False
    chunk_strategy: Optional[str] = "none"
    estimated_chunks: Optional[int] = 1
    chunk_manifest: Optional[Dict[str, Any]] = None  # populated by chunker step


class ProcessStepDef(BaseModel):
    step_id: Optional[str] = None
    step_name: str
    summary: Optional[str] = None
    details: Optional[str] = None
    expected_output: Optional[Dict[str, Any]] = None
    sub_steps: Optional[List[Dict[str, Any]]] = None
    instructional_sub_steps: Optional[List[Dict[str, Any]]] = None


class ExtractionRequest(BaseModel):
    files: List[FileEntry]
    process_step: ProcessStepDef
    documents_folder: Optional[str] = "/documents"
    output_dir: Optional[str] = None
    process_id: Optional[str] = None   # e.g. "proc-extract-asset-spreadsheet"


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "service": "document-extractor", "model": LLM_MODEL}


class RawTextRequest(BaseModel):
    filepaths: List[str]
    documents_folder: Optional[str] = "/documents"


@app.post("/raw_text")
def extract_raw_text(request: RawTextRequest):
    """
    Return full page-by-page raw text from one or more PDFs.
    Used by the orchestrator to feed source content to the step validator.
    """
    results = {}
    for fp in request.filepaths:
        path = fp if Path(fp).is_absolute() else str(Path(request.documents_folder) / fp)
        filename = Path(path).name
        try:
            with pdfplumber.open(path) as pdf:
                pages = []
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    raw_tables = page.extract_tables() or []
                    page_tables = []
                    for tbl in raw_tables:
                        if tbl and len(tbl) > 1:
                            page_tables.append({
                                "headers": [str(h or "") for h in tbl[0]],
                                "rows":    [[str(c or "") for c in row] for row in tbl[1:]],
                            })
                    pages.append({"page": page.page_number, "text": text, "tables": page_tables})
            full_text = "\n\n".join(
                f"--- Page {p['page']} ---\n{p['text']}" for p in pages if p["text"].strip()
            )
            results[filename] = {
                "filepath": path, "total_pages": len(pages),
                "pages": pages, "full_text": full_text, "error": None,
            }
        except Exception as e:
            results[filename] = {
                "filepath": path, "total_pages": 0,
                "pages": [], "full_text": "", "error": str(e),
            }
    return {"documents": results}


@app.post("/extract")
def extract_documents(request: ExtractionRequest):
    """
    Extract and tag content from each file according to the process step.
    Returns a consolidated extraction report and saves per-file JSON files.
    """
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    extractions = []

    # Route _llm_call to the appropriate model for this process
    if request.process_id == "proc-extract-asset-spreadsheet":
        _req_context.model = ASSET_EXTRACTION_MODEL
        logger.info(f"Asset extraction — using model: {ASSET_EXTRACTION_MODEL}")
    else:
        _req_context.model = LLM_MODEL

    # Use job-scoped output directory when provided, else fall back to module-level OUTPUT_DIR
    eff_output_dir = Path(request.output_dir) if request.output_dir else OUTPUT_DIR
    try:
        eff_output_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        eff_output_dir = OUTPUT_DIR
        eff_output_dir.mkdir(parents=True, exist_ok=True)

    for fe in request.files:
        filepath = fe.filepath or str(Path(request.documents_folder) / fe.filename)
        if request.process_id == "proc-extract-asset-spreadsheet":
            logger.info(f"Asset extraction file: filename={fe.filename!r}, filepath={fe.filepath!r}, "
                        f"documents_folder={request.documents_folder!r}, resolved={filepath!r}")
            result = _extract_asset_spreadsheet(fe, filepath)
        else:
            result = _extract_file(fe, filepath, request.process_step)

        # Save per-file output — include step slug so concurrent steps on the same file don't collide
        safe_name  = re.sub(r"[^\w]", "_", fe.filename)
        _step_id   = (request.process_step.step_id or request.process_step.step_name or "") if request.process_step else ""
        _step_slug = re.sub(r"[^\w]", "_", _step_id)[:30] if _step_id else ""
        _file_sfx  = f"_{_step_slug}" if _step_slug else ""
        out_path   = eff_output_dir / f"Extraction_{safe_name}{_file_sfx}_{timestamp}.json"
        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)
        result["output_file"] = str(out_path)

        extractions.append(result)

    # Consolidated report — include process_id slug so parallel steps don't collide
    _proc_slug  = re.sub(r"[^\w]", "_", request.process_id or "")[:40] if request.process_id else ""
    _report_sfx = f"_{_proc_slug}" if _proc_slug else ""
    report_path = eff_output_dir / f"ExtractionReport{_report_sfx}_{timestamp}.json"
    report = {
        "process_step": request.process_step.dict(),
        "documents_folder": request.documents_folder,
        "total_files": len(extractions),
        "extractions": extractions,
        "timestamp": timestamp,
        "output_file": str(report_path),
    }
    with open(report_path, "w") as f:
        json.dump(report, f, indent=2)

    # For asset spreadsheet extraction, also produce a flat AssetExtract JSON
    # consolidating all asset_records across all files for use in downstream steps.
    if request.process_id == "proc-extract-asset-spreadsheet":
        all_asset_records = []
        source_documents  = []
        for ex in extractions:
            se = ex.get("step_extraction") or {}
            records = se.get("asset_records") or []
            all_asset_records.extend(records)
            if ex.get("filename"):
                source_documents.append(ex["filename"])
        asset_extract_path = eff_output_dir / f"AssetExtract_{timestamp}.json"
        asset_extract = {
            "process_id":       "proc-extract-asset-spreadsheet",
            "timestamp":        timestamp,
            "source_documents": source_documents,
            "total_assets":     len(all_asset_records),
            "asset_records":    all_asset_records,
            "output_file":      str(asset_extract_path),
        }
        with open(asset_extract_path, "w") as f:
            json.dump(asset_extract, f, indent=2)
        report["asset_extract"]      = asset_extract
        report["asset_extract_file"] = str(asset_extract_path)
        logger.info(f"AssetExtract saved: {asset_extract_path} ({len(all_asset_records)} records)")

    return report


# ── Core extraction logic ─────────────────────────────────────────────────────

def _extract_file(fe: FileEntry, filepath: str, step: ProcessStepDef) -> dict:
    """Full extraction pipeline for a single document."""
    # If chunk PNGs are available, use vision-based extraction instead of PDF text parsing
    if fe.chunk_manifest and fe.chunk_manifest.get("chunks"):
        return _extract_chunked_file(fe, fe.chunk_manifest, step)

    tool_id = (fe.processing_tool_id or "").lower()

    # Determine page grouping for large documents.
    # For quadrant-split (A2/A1/A0 drawings), use group_size=1 so every page
    # becomes its own chunk — prevents large drawing pages from being merged.
    # For page-split, same treatment. For none, use default (let detector decide).
    chunk_override: Optional[int] = None
    if fe.requires_chunking:
        if fe.chunk_strategy in ("quadrant-split", "page-split"):
            chunk_override = 1  # one page per chunk

    # 1. Extract raw sections from the document
    if filepath.lower().endswith(".pdf") or "pdf" in tool_id:
        raw_sections, doc_quality = _extract_pdf_sections(filepath, chunk_page_size=chunk_override)
    else:
        raw_sections = _extract_text_file(filepath)
        doc_quality = "text"

    # 2. Tag all sections in batched LLM calls (chunked)
    tagged_sections = _tag_all_sections_chunked(raw_sections, step)

    # 3. Isolate relevant sections (score >= 0.4)
    relevant = [s for s in tagged_sections if s.get("relevance_score", 0) >= 0.4]

    base = {
        "filename":              fe.filename,
        "filepath":              filepath,
        "document_type":         fe.document_type,
        "document_category":     fe.document_category,
        "input_text_quality":    fe.text_quality,   # quality as assessed by planner
        "document_quality":      doc_quality,       # quality as computed from PDF content
        "page_size":             fe.page_size or "A4",
        "requires_chunking":     fe.requires_chunking or False,
        "chunk_strategy":        fe.chunk_strategy or "none",
        "estimated_chunks":      fe.estimated_chunks or 1,
        "process_step_name":     step.step_name,
        "total_sections":        len(raw_sections),
        "relevant_sections":     len(relevant),
    }

    # 4a. If sub_steps defined, run a targeted extraction per sub-step.
    # Pass all tagged sections so each sub-step can apply its own threshold.
    if step.sub_steps:
        base["sub_step_extractions"] = _extract_per_sub_step(tagged_sections, step.sub_steps)
    else:
        # 4b. Single structured extraction across all relevant sections
        base["step_extraction"] = _structured_extraction(relevant, step)

    return base


_MARKER_WORDS = ["ATTENTION", "WARNING", "CAUTION", "NOTICE"]

# Numbered note pattern — matches "1. text", "2. text" etc at start of line or after newline
_NOTE_NUM_RE = re.compile(r"(?:^|\n)\s*(\d{1,2})\.\s+\S", re.MULTILINE)


def _find_note_numbers(text: str) -> set:
    """Return set of note numbers (ints 1-25) found in text."""
    return {int(m.group(1)) for m in _NOTE_NUM_RE.finditer(text) if int(m.group(1)) <= 25}


def _verify_and_fill_extraction(
    consolidated: dict,
    combined_raw: str,
    sub_step_ids: List[str],
    sub_step_schema_lines: List[str],
) -> dict:
    """Phase 4 — Programmatic gap detection followed by a targeted LLM fill pass.

    Checks each sub-step extraction against the full raw text corpus for:
    1. Missing numbered notes (gaps in the 1..N sequence found in raw text)
    2. Missing marker-word callouts (ATTENTION / WARNING / CAUTION / NOTICE) that
       appear as standalone annotations in the raw text but not in the extraction
    3. Detected gaps trigger a single focused LLM re-pass that corrects only the
       affected fields — complete fields are not changed.

    Returns the (possibly corrected) consolidated dict.
    """
    if not isinstance(consolidated, dict):
        return consolidated

    raw_upper        = combined_raw.upper()
    raw_note_nums    = _find_note_numbers(combined_raw)
    raw_markers      = [m for m in _MARKER_WORDS if m in raw_upper]

    gaps: List[str] = []
    affected_ids: List[str] = []

    for sid in sub_step_ids:
        val = consolidated.get(sid)
        if not isinstance(val, str) or not val.strip():
            continue

        val_upper = val.upper()
        field_gaps: List[str] = []

        if "note" in sid.lower():
            # ── Check 1: numbered note sequence ───────────────────────────────
            extracted_nums  = _find_note_numbers(val)
            missing_nums    = raw_note_nums - extracted_nums
            if missing_nums:
                field_gaps.append(
                    f"Missing numbered note(s) {sorted(missing_nums)} "
                    f"(present in raw text, absent from extraction)"
                )

            # ── Check 2: marker-word callouts ─────────────────────────────────
            # A marker in raw text that does NOT appear in the extraction at all
            for marker in raw_markers:
                if marker not in val_upper:
                    field_gaps.append(
                        f"Marker word '{marker}:' is present in raw text as a standalone "
                        f"callout annotation but is not captured in this field"
                    )

            # ── Check 3: truncated sentences ──────────────────────────────────
            # Last non-empty line ends mid-word (no punctuation and short word)
            last_line = val.rstrip().split("\n")[-1].strip()
            if last_line and not last_line[-1] in ".,:;)\"'" and len(last_line.split()) <= 3:
                field_gaps.append(
                    f"Extraction appears to end mid-sentence: '...{last_line}'"
                )

        if field_gaps:
            gaps.extend(field_gaps)
            affected_ids.append(sid)
            for g in field_gaps:
                logger.info(f"[Verification] gap in '{sid}': {g}")

    if not gaps:
        logger.info("[Verification] No gaps detected — extraction is complete")
        return consolidated

    logger.info(f"[Verification] {len(gaps)} gap(s) across {len(affected_ids)} field(s) — running gap-fill pass")

    # Build a targeted fill prompt — only ask for the affected fields
    affected_schema = "\n".join(
        line for line in sub_step_schema_lines
        if any(sid in line for sid in affected_ids)
    )
    affected_keys = ", ".join(f'"{sid}"' for sid in affected_ids)
    current_values = {sid: consolidated.get(sid) for sid in affected_ids}

    gap_fill_prompt = (
        f"EXTRACTION REVIEW AND COMPLETION TASK\n\n"
        f"The following gaps were detected in a prior extraction pass:\n" +
        "\n".join(f"  - {g}" for g in gaps) +
        f"\n\nFULL RAW TEXT EXTRACTED FROM ALL DOCUMENT CHUNKS (in spatial order):\n"
        f"{combined_raw[:12000]}\n\n"
        f"CURRENT (INCOMPLETE) VALUES FOR AFFECTED FIELDS:\n"
        f"{json.dumps(current_values, indent=2)[:3000]}\n\n"
        f"TASK: Review the raw text carefully and produce COMPLETE, corrected values "
        f"for ONLY the following fields:\n{affected_schema}\n\n"
        f"Gap-fill rules:\n"
        f"- Include ALL numbered notes from the raw text in sequence order (1., 2., 3., ...)\n"
        f"- Include ALL standalone callout annotations: any text preceded by ATTENTION:, "
        f"WARNING:, CAUTION:, or NOTICE: — even if they repeat content from the numbered notes, "
        f"capture them as separate entries appended after the numbered list\n"
        f"- Reconstruct split text: the raw text has left-half and right-half fragments from "
        f"adjacent image columns — join them into complete sentences\n"
        f"- Do NOT truncate — if a note or callout seems to continue, find and include the rest "
        f"from the raw text\n"
        f"- Return ONLY the fields listed above, as a JSON object\n"
        f"- Do NOT wrap output in markdown fences — return raw JSON only\n"
        f"Required keys: {affected_keys}"
    )

    try:
        filled = _llm_call(gap_fill_prompt)
        if isinstance(filled, dict):
            for sid in affected_ids:
                if sid in filled and filled[sid] is not None:
                    consolidated[sid] = filled[sid]
                    logger.info(f"[Verification] gap-filled '{sid}': {str(filled[sid])[:80]}...")
        logger.info("[Verification] gap-fill pass complete")
    except Exception as e:
        logger.warning(f"[Verification] gap-fill pass failed: {e}")

    return consolidated


def _stitch_adjacent_chunks(
    chunks: List[dict],
    chunk_results: List[dict],
    sub_step_ids: List[str],
) -> List[dict]:
    """Phase 1.5 — For adjacent chunk pairs where content is cut off at a boundary,
    run a multi-image Gemini call with both chunks visible simultaneously to assemble
    the complete spanning text.  Only pairs that carry continuation markers (→ / ←)
    in their raw_text are processed; all others are skipped to limit LLM cost.

    Returns a list of stitch records:
      {"chunk_ids": [...], "stitched_raw": "...", "data": {...}}
    """
    # Build a (chunk_meta, chunk_result) list sorted by page then sequence
    ordered: List[tuple] = []
    result_by_id = {cr.get("chunk_id"): cr for cr in chunk_results}
    for chunk in sorted(chunks, key=lambda c: (c.get("page_number", 1), c.get("sequence", 0))):
        cr = result_by_id.get(chunk.get("chunk_id"))
        if cr and cr.get("extracted"):
            ordered.append((chunk, cr))

    stitched: List[dict] = []

    for i in range(len(ordered) - 1):
        chunk_a, result_a = ordered[i]
        chunk_b, result_b = ordered[i + 1]

        # Only stitch chunks on the same page
        if chunk_a.get("page_number") != chunk_b.get("page_number"):
            continue

        raw_a = (result_a.get("extracted") or {}).get("raw_text", "") or ""
        raw_b = (result_b.get("extracted") or {}).get("raw_text", "") or ""

        # Only stitch when a continuation marker is present at this boundary
        if "\u2192" not in raw_a and "\u2190" not in raw_b:
            continue

        fpath_a = chunk_a.get("filepath", "")
        fpath_b = chunk_b.get("filepath", "")
        if not (fpath_a and fpath_b and Path(fpath_a).exists() and Path(fpath_b).exists()):
            logger.warning(f"Skipping stitch: missing file for chunks {chunk_a.get('chunk_id')}/{chunk_b.get('chunk_id')}")
            continue

        schema = (
            "{\n" +
            ",\n".join(f'  "{sid}": assembled value or null' for sid in sub_step_ids) +
            "\n}"
        )
        stitch_prompt = (
            f"You are an expert engineering document analyst.\n"
            f"The two images shown are ADJACENT regions of the same engineering drawing page.\n"
            f"Image 1 covers region '{chunk_a.get('region')}'. "
            f"Image 2 covers region '{chunk_b.get('region')}' (immediately to the right of or below Image 1).\n\n"
            f"Known partial text from Image 1 (last 800 chars, may be cut off at the boundary):\n"
            f"{raw_a[-800:]}\n\n"
            f"Known partial text from Image 2 (first 800 chars, may continue from Image 1):\n"
            f"{raw_b[:800]}\n\n"
            f"TASK: Identify any text, numbered notes, or data entries that are split across the "
            f"boundary between these two images. Assemble the complete, uninterrupted text for each "
            f"such entry by reading both images together.\n\n"
            f"Return a JSON object with exactly two keys:\n"
            f"  'stitched_raw': a single string containing the complete assembled text for any "
            f"boundary-spanning content (empty string if nothing spans the boundary),\n"
            f"  'data': {schema}\n"
            f"Only populate 'data' fields where having both images gives you MORE COMPLETE information "
            f"than either image alone. Use null for fields not affected by this boundary."
        )

        try:
            img_a = Path(fpath_a).read_bytes()
            img_b = Path(fpath_b).read_bytes()
            result = _llm_call_vision_multi(stitch_prompt, [img_a, img_b])
            stitched_raw = result.get("stitched_raw", "") if isinstance(result, dict) else ""
            stitched_data = result.get("data", {}) if isinstance(result, dict) else {}
            logger.info(
                f"Boundary stitch {chunk_a.get('chunk_id')}/{chunk_b.get('chunk_id')}: "
                f"{len(stitched_raw)} chars assembled"
            )
            stitched.append({
                "chunk_ids":    [chunk_a.get("chunk_id"), chunk_b.get("chunk_id")],
                "stitched_raw": stitched_raw,
                "data":         stitched_data if isinstance(stitched_data, dict) else {},
            })
        except Exception as e:
            logger.warning(
                f"Boundary stitch failed for {chunk_a.get('chunk_id')}/{chunk_b.get('chunk_id')}: {e}"
            )

    return stitched


def _extract_chunked_file(fe: FileEntry, chunk_manifest: dict, step: ProcessStepDef) -> dict:
    """
    Vision-based extraction for documents that have been pre-split into PNG chunks.

    Each chunk PNG is sent to Gemini with the process step instructions.
    Results from all chunks are consolidated into a single structured output
    that matches the format produced by _extract_file.
    """
    chunks     = chunk_manifest.get("chunks", [])
    strategy   = chunk_manifest.get("chunk_strategy", fe.chunk_strategy or "quadrant-split")
    page_size  = chunk_manifest.get("page_size", fe.page_size or "A4")
    total_pages = chunk_manifest.get("total_pages", 1)

    # Build sub-step schema split by phase.
    # Phase 1 sub-steps are extracted from image chunks (vision).
    # Phase 2 sub-steps are analysed from the consolidated Phase 1 text (text-only LLM).
    # Sub-steps without an explicit "phase" field default to Phase 1.
    _p1_ids: List[str] = []   # Phase 1 — image-based extraction
    _p1_schema: List[str] = []
    _p2_ids: List[str] = []   # Phase 2 — text-based analysis of Phase 1 output
    _p2_schema: List[str] = []
    _p2_steps: List[dict] = []  # Full sub-step dicts for Phase 2 (used in analysis prompt)
    _p1_output_formats: dict = {}  # sid -> output_format string for Phase 1 sub-steps

    if step.sub_steps:
        for ss in step.sub_steps:
            ss_id    = ss.get("sub_step_id") or f"sub_step_{ss.get('sub_step_number', '')}"
            ss_name  = ss.get("sub_step_name", ss_id)
            ss_det   = (ss.get("details") or "")[:200]
            ss_phase = ss.get("phase", 1)
            schema_line = f'  "{ss_id}": <{ss_name}: {ss_det}>'
            if ss_phase == 2:
                _p2_ids.append(ss_id)
                _p2_schema.append(schema_line)
                _p2_steps.append(ss)
            else:
                _p1_ids.append(ss_id)
                _p1_schema.append(schema_line)
                if ss.get("output_format"):
                    _p1_output_formats[ss_id] = ss["output_format"]

    # Backward-compatible aliases — used by existing code below
    _sub_step_ids          = _p1_ids
    _sub_step_schema_lines = _p1_schema

    # Sub-steps whose output_format declares a JSON array — need special handling
    # (accumulation across chunks, positional label pairing). Defined here so Phase 1.25 can use it.
    _array_sub_steps: set = {
        sid for sid in _p1_ids
        if sid in _p1_output_formats and _p1_output_formats[sid].lstrip().startswith("JSON array")
    }

    # Build extraction instruction from the process step definition
    inst_parts = [f"Process step: {step.step_name}"]
    if step.details:
        inst_parts.append(f"Details: {step.details[:300]}")
    if _sub_step_schema_lines:
        inst_parts.append("Fields to extract:")
        inst_parts.extend(_sub_step_schema_lines)
    instruction = "\n".join(inst_parts)

    # ── Page-level legend presence check (pre-Phase 1) ────────────────────────
    # For array sub-steps (e.g. legend extraction), determine which pages actually
    # contain a legend before running per-chunk extraction. Each page's chunks are
    # sent together as a single multi-image call so the model sees the full page context.
    # Pages with no legend will have their array sub-step entries discarded after Phase 1.
    pages_with_legend: set = set(range(1, total_pages + 1))  # conservative default: all pages
    if _array_sub_steps:
        chunks_by_page: dict = {}
        for c in chunks:
            chunks_by_page.setdefault(c.get("page_number", 1), []).append(c)
        confirmed_legend_pages: set = set()
        for pg, page_chunks in sorted(chunks_by_page.items()):
            sorted_pc = sorted(page_chunks, key=lambda c: c.get("sequence", 0))
            # Use all chunks so the model sees the entire page; cap at 12 to stay within limits
            sample = sorted_pc[:12]
            valid = [c for c in sample if c.get("filepath") and Path(c["filepath"]).exists()]
            if not valid:
                logger.warning(f"Page {pg}: no valid chunk images for legend check — assuming present")
                confirmed_legend_pages.add(pg)
                continue
            try:
                img_bytes_list = [Path(c["filepath"]).read_bytes() for c in valid]
                if _vision_page_has_legend(img_bytes_list):
                    confirmed_legend_pages.add(pg)
                    logger.info(f"Page {pg}: legend confirmed — array sub-steps will be extracted")
                else:
                    logger.info(f"Page {pg}: no legend detected — array sub-step entries will be discarded")
            except Exception as e:
                logger.warning(f"Page {pg} legend check error: {e} — assuming present")
                confirmed_legend_pages.add(pg)
        if confirmed_legend_pages:
            pages_with_legend = confirmed_legend_pages

    # ── Phase 1: process each chunk with the vision model ─────────────────────
    # Sort chunks into reading order (page asc, then sequence asc) so the prompt
    # context and boundary markers are coherent.
    chunks = sorted(chunks, key=lambda c: (c.get("page_number", 1), c.get("sequence", 0)))

    chunk_results: List[dict] = []
    for chunk in chunks:
        chunk_id   = chunk.get("chunk_id", "?")
        region     = chunk.get("region", "")
        page_num   = chunk.get("page_number", 1)
        fpath      = chunk.get("filepath", "")

        if not fpath or not Path(fpath).exists():
            logger.warning(f"Chunk file missing: {fpath}")
            chunk_results.append({"chunk_id": chunk_id, "region": region,
                                   "page_number": page_num, "extracted": None,
                                   "error": "file not found"})
            continue

        if _sub_step_ids:
            schema_lines = []
            for sid in _sub_step_ids:
                if sid in _p1_output_formats:
                    # Use the first line of output_format as the inline value hint
                    fmt_hint = _p1_output_formats[sid].split("\n")[0][:200]
                    schema_lines.append(f'    "{sid}": {fmt_hint}')
                else:
                    schema_lines.append(f'    "{sid}": extracted value or null')
            data_schema = "{\n" + ",\n".join(schema_lines) + "\n  }"
        else:
            data_schema = "relevant extracted fields or null"

        prompt = (
            f"You are an expert engineering document analyst.\n"
            f"This image is region '{region}' (chunk {chunk_id}, page {page_num} of {total_pages}) "
            f"from a {page_size} engineering drawing split using {strategy}.\n\n"
            f"{instruction}\n\n"
            f"BOUNDARY AWARENESS: This image is one tile of a grid. Text may be cut off at the edges.\n"
            f"- If text is cut off at the RIGHT or BOTTOM edge (continues in the next tile), "
            f"append \u2192 to that line in raw_text.\n"
            f"- If text starts abruptly at the LEFT or TOP edge (continues from a previous tile), "
            f"prepend \u2190 to that line in raw_text.\n"
            f"- Always capture partially-cut-off text in raw_text — do NOT omit it just because "
            f"it is incomplete at the boundary.\n\n"
            f"LEGEND EXTRACTION RULE: If this chunk contains a legend or symbol table, each legend row "
            f"has a DRAWN SYMBOL on the left and a TEXT LABEL on the right (or directly below) the symbol. "
            f"You MUST read BOTH the visual symbol AND the text label for every row and include them "
            f"together in each legend entry. Do not leave the label field empty if label text is "
            f"visible anywhere in this image next to or near the symbol.\n\n"
            f"Return a JSON object with exactly two keys:\n"
            f"  'raw_text': all legible text visible in this image chunk (preserve as plain text, "
            f"  use \u2192/\u2190 markers at cut-off boundaries as instructed above),\n"
            f"  'data': {data_schema}\n"
            f"Use null for any field not visible in this specific chunk. "
            f"Only report information actually present in this image."
        )

        try:
            img_bytes = Path(fpath).read_bytes()
            result    = _llm_call_vision(prompt, img_bytes)
            chunk_results.append({
                "chunk_id":    chunk_id,
                "sequence":    chunk.get("sequence", 0),
                "page_number": page_num,
                "region":      region,
                "extracted":   result,
            })
        except Exception as e:
            logger.warning(f"Vision extraction failed for {chunk_id}: {e}")
            chunk_results.append({
                "chunk_id":      chunk_id,
                "sequence":      chunk.get("sequence", 0),
                "page_number":   page_num,
                "region":        region,
                "extracted":     None,
                "skipped":       True,
                "skip_reason":   str(e),
            })

    # ── Phase 1.1: quadrant retry for full-page chunks with empty legend ────────
    # Large-format drawings (A3/A1) are chunked as a single full-page image.
    # The vision model often misses the legend when it's a small table in one
    # corner of a dense engineering drawing.  When Phase 1 returned no legend
    # entries for a full-page chunk, crop the image into four quadrants (TL, TR,
    # BL, BR) and run each through the same Phase 1 prompt.  Only entries for
    # array sub-steps are collected; raw_text is merged back.  Quadrant results
    # are appended to chunk_results with synthetic chunk IDs so Phase 1.25 and
    # the merge logic treats them like any other chunk.
    if _array_sub_steps:
        _fullpage_empty = [
            cr for cr in chunk_results
            if cr.get("region") == "full-page"
            and not cr.get("skipped")
            and cr.get("extracted") is not None
            and not any(
                cr["extracted"].get("data", {}).get(sid)
                for sid in _array_sub_steps
            )
        ]
        for _fp_cr in _fullpage_empty:
            _fp_path = None
            # find the filepath from the original chunks list
            for _oc in chunks:
                if _oc.get("chunk_id") == _fp_cr.get("chunk_id"):
                    _fp_path = _oc.get("filepath")
                    break
            if not _fp_path or not Path(_fp_path).exists():
                continue
            try:
                _img = PIL.Image.open(_fp_path)
                _w, _h = _img.size
                _quadrants = {
                    "quad-tl": (0,       0,       _w // 2, _h // 2),
                    "quad-tr": (_w // 2, 0,       _w,      _h // 2),
                    "quad-bl": (0,       _h // 2, _w // 2, _h),
                    "quad-br": (_w // 2, _h // 2, _w,      _h),
                }
                logger.info(
                    f"Phase 1.1: full-page chunk {_fp_cr['chunk_id']} had no legend entries — "
                    f"retrying with 4 quadrant crops ({_w}×{_h}px)"
                )
                _quad_found = 0
                for _qname, _box in _quadrants.items():
                    _qimg = _img.crop(_box)
                    _qbuf = io.BytesIO()
                    _qimg.save(_qbuf, format="PNG")
                    _qbytes = _qbuf.getvalue()
                    _qprompt = (
                        f"You are an expert engineering document analyst.\n"
                        f"This image is a {_qname.replace('quad-', '').upper()} quadrant crop "
                        f"of a {page_size} engineering drawing.\n\n"
                        f"{instruction}\n\n"
                        f"LEGEND EXTRACTION RULE: If this quadrant contains a legend or symbol "
                        f"table, each legend row has a DRAWN SYMBOL on the left and a TEXT LABEL "
                        f"on the right (or directly below) the symbol. "
                        f"You MUST read BOTH the visual symbol AND the text label for every row. "
                        f"Engineering drawing legends are typically located in the bottom-right "
                        f"or bottom-left corner of the sheet.\n\n"
                        f"Return a JSON object with exactly two keys:\n"
                        f"  'raw_text': all legible text visible in this quadrant,\n"
                        f"  'data': {data_schema}\n"
                        f"Use null for any field not visible in this quadrant. "
                        f"Only report information actually present in this image."
                    )
                    try:
                        _qresult = _llm_call_vision(_qprompt, _qbytes)
                        _qentries = (_qresult.get("data") or {}) if isinstance(_qresult, dict) else {}
                        _has_entries = any(
                            isinstance(_qentries.get(sid), list) and _qentries[sid]
                            for sid in _array_sub_steps
                        )
                        chunk_results.append({
                            "chunk_id":    f"{_fp_cr['chunk_id']}-{_qname}",
                            "sequence":    _fp_cr.get("sequence", 0) + 0.1,
                            "page_number": _fp_cr.get("page_number", 1),
                            "region":      _qname,
                            "extracted":   _qresult,
                        })
                        if _has_entries:
                            _quad_found += 1
                            logger.info(f"Phase 1.1: {_qname} yielded legend entries")
                    except Exception as _qe:
                        logger.warning(f"Phase 1.1 quadrant {_qname} failed: {_qe}")
                logger.info(
                    f"Phase 1.1: quadrant retry complete — {_quad_found}/4 quadrants yielded entries"
                )
            except Exception as _e:
                logger.warning(f"Phase 1.1: could not crop {_fp_path}: {_e}")

    # ── Phase 1.25: positional label pairing for array sub-steps ─────────────
    # Engineering legend tables have symbols on the LEFT and label text on the RIGHT.
    # The vision model correctly identifies the symbol geometry but often fails to pair
    # the adjacent label text. This pass uses each chunk's own raw_text as a label source
    # and matches label candidates to symbol entries positionally (same sequential order).
    # Patterns that mark the END of the legend — stop extracting candidates here
    _LEGEND_SECTION_END = re.compile(
        r'TEMPLATE VERSION|MAY NOT BE COPIED|COPYRIGHT THEREIN|WRITTEN CONSENT'
        r'|PROPERTY OF ACME|REPRODUCED[,.]|DISTRIBUTED[,.]|LOANED OR'
        r'|REFERENCE DRAWING|CONNECTION OF LOAD',
        re.IGNORECASE,
    )
    # Noise lines that look uppercase but are not legend labels
    _LEGEND_NOISE = re.compile(
        r'^[\d\s.,()/:@+\-]+$'                                  # pure numbers/punctuation
        r'|©|www\.|http|\+61|@'                                 # web/email markers
        r'|ORIGINAL\s+ISSUE|DRAFT\s+No|AMENDMENT'              # title block history
        r'|SCALE\s*\d|\bDATE\s+\d|\bDRAWN\s+BY\b'             # drawing metadata
        r'|CADASTRE|LAND AND PROPERTY'                          # attribution
        r'|TELEPHONE\s*:|FACSIMILE\s*:|GPO\s+BOX'             # contact details
        r'|GEORGE\s+STREET|ERNST\s+&|YOUNG\s+CENTRE'           # address
        r'|ASP\s+REF|ACCREDIATION|ACCREDITATION'               # accreditation
        r'|\bSITE\s+PLAN\s*$|\bGENERAL\s*$|\bOVERHEAD\s*$'   # standalone section labels
        r'|\bUNDERGROUND\s*$|\bSUBSTATIONS\s*$',              # in work-order table
        re.IGNORECASE,
    )
    # Common single English words that are never legend labels
    _NON_LABELS = {
        "AND", "THE", "OF", "OR", "FOR", "TO", "IN", "AT", "BY", "AS",
        "IS", "IT", "BE", "AN", "ON", "NO", "SO", "DO", "UP",
        "BUT", "NOT", "ARE", "HAS", "HAD", "WAS", "THIS", "WITH",
    }
    # ── Normalize label_text → label and filter bad entries across all chunks ──
    # The vision model sometimes returns the label in 'label_text' instead of 'label',
    # or has a shifted 'label' while 'label_text' is correct. Prefer 'label_text' when present.
    # Also filter obviously non-legend entries (document references, very long text, etc.)
    _BAD_LABEL_RE = re.compile(
        r'\bFPJ\d{4}\b|\bFPJ-\d{4}\b'                     # work order references
        r'|AGREEMENT FOR ENTRY|ENVIRONMENTAL REPORT'        # document references
        r'|EQUIPMENT TO BE RETURNED|TELECOMMUNICATION ASSET'
        r'|SER\s*\)|GRANT AND CREATION'
        r'|DESIGNER.S SAFETY|ABORIGINAL CULTURAL|CULTURAL HERITAGE'
        r'|DISPENSATION FORM|SKETCH SR\s*\d|LETTER FOR CONFIRMATION'
        r'|SAFETY REPORT|ASP LEVEL\s+\d|ACCREDITATION|LEVEL 2/'
        r'|^OVERHEAD$|^UNDERGROUND$|^SUBSTATIONS?$|^GENERAL$'
        r'|ORIGINAL\s+SCALE|ORIGINAL\s+ISSUE|AMENDMENT\s+\d'
        r'|\bLGA\b(?!\s+DEMARCATION)|DUCT USAGE CHARGES|(?<![A-Za-z])AVOUR ENERGY'
        r'|^NIL$|^-\s*NIL\b|^-\s+[A-Z]|PM SUB\s+\d+|ESTABLISHED ON ARP\d+'
        r'|\bFUNDED\b|\bCONTESTABLE\b|\bCUSTOMER\b(?!\s+CONNECT)'
        r'|CO-ORDINATION SUPPLY|TO BE CONFIRMED|Third Quarter'
        r'|RETURNED TO NEAREST|CONDUCTOR\s+[A-Z]\s+\'',
        re.IGNORECASE,
    )
    # Canonical symbol descriptions and categories from the drawing reference table.
    # Applied programmatically after accumulation to ensure consistent descriptions.
    _SYMBOL_REF: dict = {
        "NEW LV TRENCH":               ("Long dashed line",                                              "cable"),
        "STRING NEW OH CABLE":         ("Short dashed line",                                             "cable"),
        "EXISTING UNDERGROUND MAINS":  ("Alternating long-short dashed line",                           "cable"),
        "EXISTING OH CABLE":           ("Solid thin line",                                               "cable"),
        "REMOVE CONDUCTOR":            ("Dotted line",                                                   "cable"),
        "EXISTING DUCTS":              ("Thick heavy solid black line",                                  "cable"),
        "NEW HV TRENCH":               ("Solid line with two diagonal slash marks",                      "cable"),
        "EXISTING LANTERN":            ("Small circle with internal cross (four-quadrant cross inside circle)", "equipment"),
        "REMOVE LANTERN":              ("Circle with internal cross and starburst outer spikes radiating from edge", "equipment"),
        "NEW LANTERN":                 ("Circle with starburst outer spikes and hollow centre",          "equipment"),
        "EXISTING POLE":               ("Small solid black filled circle",                               "equipment"),
        "REMOVE POLE":                 ("Small solid black filled circle with large X through it",       "equipment"),
        "REPLACE POLE":                ("Circle split vertically half-black half-white",                 "equipment"),
        "NEW POLE":                    ("Small hollow open circle",                                      "equipment"),
        "EXISTING COLUMN":             ("Small solid black filled square",                               "equipment"),
        "NEW COLUMN":                  ("Small hollow open square",                                      "equipment"),
        "EXISTING PILLAR":             ("Solid black filled rectangle",                                  "equipment"),
        "NEW PILLAR":                  ("Hollow rectangle outline",                                      "equipment"),
        "PADMOUNT SUBSTATION":         ("Rectangle containing two triangles touching at their points",   "substation"),
        "POLE SUBSTATION":             ("Circle with triangle inside",                                   "substation"),
        "LV LINK (N/O)":              ("Small circle with short vertical ticks on top and bottom",      "equipment"),
        "HV ABS (N/C)":               ("Circle with horizontal line through the centre",                "equipment"),
        "HV USL (N/C)":               ("Semi-circle dome with horizontal base line",                    "equipment"),
        "HIGH PRESSURE GAS":           ("Solid line with GAS text centred along it",                    "other"),
        "WATER MAIN":                  ("Solid line with W characters spaced evenly along it",           "other"),
        "LGA DEMARCATION":             ("Thin solid line",                                               "boundary"),
        "NEW FREEWAY BOUNDARY":        ("Faint light grey long dashed line",                             "boundary"),
        "ACME ENERGY EASEMENT":   ("Faint light grey medium dashed line",                           "boundary"),
    }
    for cr in chunk_results:
        ext = cr.get("extracted")
        if not isinstance(ext, dict):
            continue
        data = ext.get("data") or {}
        for sid in _array_sub_steps:
            entries = data.get(sid)
            if not isinstance(entries, list):
                continue
            cleaned: list = []
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                lt  = (entry.get("label_text") or "").strip()
                lbl = (entry.get("label") or "").strip()
                # Prefer label_text over label (fixes off-by-one shifts from vision model)
                if lt:
                    # Normalise newlines inside label_text (multi-line cells in some chunk outputs)
                    lt = re.sub(r'\s*\n\s*', ' ', lt).strip()
                    entry["label"] = lt
                    lbl = lt
                elif lbl:
                    # Clean up newlines in label too
                    lbl = re.sub(r'\s*\n\s*', ' ', lbl).strip()
                    entry["label"] = lbl
                # Drop entries whose label looks like a document reference, not a legend item
                if lbl and (_BAD_LABEL_RE.search(lbl) or len(lbl) > 70):
                    continue
                # Drop entries with no symbol and no label (completely empty)
                sym = (entry.get("symbol_description") or "").strip()
                if not lbl and not sym:
                    continue
                cleaned.append(entry)
            data[sid] = cleaned

    for cr in chunk_results:
        ext = cr.get("extracted")
        if not isinstance(ext, dict):
            continue
        data = ext.get("data") or {}
        raw_text = ext.get("raw_text") or ""
        for sid in _array_sub_steps:
            entries = data.get(sid)
            if not isinstance(entries, list) or not entries:
                continue
            # Only repair chunks where most labels are missing
            missing = sum(1 for e in entries if isinstance(e, dict) and not (e.get("label") or "").strip())
            if missing < len(entries) * 0.5:
                continue  # Majority of labels already filled — trust the vision model
            # Extract candidate label lines from raw_text.
            # Stop as soon as we hit a copyright/section-end marker.
            candidates: List[str] = []
            for line in raw_text.split("\n"):
                s = line.strip().lstrip("←").rstrip("→").strip()
                if not s:
                    continue
                # Hard stop at copyright/legal/non-legend sections
                if _LEGEND_SECTION_END.search(s):
                    break
                # Skip legend header itself
                if re.search(r'\bLEGEND\b|\bSYMBOL\s+KEY\b', s, re.IGNORECASE):
                    continue
                if len(s) < 4 or len(s) > 60:
                    continue
                if _LEGEND_NOISE.search(s):
                    continue
                if s.upper().strip() in _NON_LABELS:
                    continue
                # Accept when ≥50% of ALPHA chars are uppercase (handles "LV LINK (N/O)")
                alpha = [c for c in s if c.isalpha()]
                if not alpha:
                    continue
                if sum(1 for c in alpha if c.isupper()) / len(alpha) >= 0.50:
                    candidates.append(s)
            if not candidates:
                continue
            # Pair if candidate count is within 40% of entry count (or within 3)
            if abs(len(candidates) - len(entries)) > max(3, len(entries) * 0.4):
                logger.debug(
                    f"Legend pairing skipped for {cr.get('chunk_id')} {sid}: "
                    f"{len(entries)} entries vs {len(candidates)} candidates"
                )
                continue
            paired = 0
            for i, entry in enumerate(entries):
                if not isinstance(entry, dict):
                    continue
                if not (entry.get("label") or "").strip() and i < len(candidates):
                    entry["label"] = candidates[i]
                    paired += 1
            if paired:
                logger.info(
                    f"Legend pairing: {cr.get('chunk_id')} {sid} — "
                    f"paired {paired} labels from raw_text"
                )

    # ── Phase 1.3: reference-label gap fill from raw text ─────────────────────
    # The vision model may miss some legend entries in Phase 1. This pass scans the raw_text
    # of chunks that already have some extracted legend entries (confirmed legend regions)
    # and injects any known reference-table labels that appear in raw_text but are absent
    # from the extracted entries for that chunk's page.
    _all_ref_labels_upper = {k.upper(): (k, v) for k, v in _SYMBOL_REF.items()}
    for cr in chunk_results:
        ext = cr.get("extracted")
        if not isinstance(ext, dict):
            continue
        data = ext.get("data") or {}
        raw_text = ext.get("raw_text") or ""
        page_num = cr.get("page_number", 1)
        raw_upper = raw_text.upper()
        for sid in _array_sub_steps:
            entries = data.get(sid, [])
            if not isinstance(entries, list) or not entries:
                continue  # Skip non-legend chunks (no entries extracted)
            # Build set of labels already in this chunk's entries
            current_labels = {
                (e.get("label") or "").strip().upper()
                for e in entries if isinstance(e, dict)
            }
            # Inject missing reference labels that appear in raw_text
            for ref_upper, (ref_lbl, (ref_sym, ref_cat)) in _all_ref_labels_upper.items():
                if ref_upper in current_labels:
                    continue
                if ref_upper in raw_upper:
                    entries.append({
                        "page": page_num,
                        "symbol_description": ref_sym,
                        "label": ref_lbl,
                        "category": ref_cat,
                    })
                    logger.info(f"Gap fill: injected '{ref_lbl}' into {cr.get('chunk_id')} {sid}")
            data[sid] = entries

    # ── Page legend filter (post Phase 1.3) ───────────────────────────────────
    # Discard array sub-step entries from pages confirmed to have no legend.
    # This removes garbage entries that the vision model extracted from title blocks,
    # cost tables, or notes on pages that don't contain a legend/symbol table.
    if _array_sub_steps:
        for cr in chunk_results:
            pg = cr.get("page_number", 1)
            if pg in pages_with_legend:
                continue
            ext = cr.get("extracted")
            if not isinstance(ext, dict):
                continue
            data = ext.get("data") or {}
            for sid in _array_sub_steps:
                if data.get(sid):
                    logger.info(
                        f"Page legend filter: clearing {len(data[sid])} {sid} entries "
                        f"from chunk {cr.get('chunk_id')} (page {pg} has no legend)"
                    )
                    data[sid] = []

    # ── Phase 1.5: boundary stitching for adjacent chunk pairs ────────────────
    # Run multi-image vision calls for chunk pairs where text is cut off at the boundary.
    # Stitched fragments supplement the per-chunk extractions with assembled spanning content.
    stitched_results: List[dict] = []
    if _sub_step_ids:
        try:
            stitched_results = _stitch_adjacent_chunks(chunks, chunk_results, _sub_step_ids)
            logger.info(f"Boundary stitching: {len(stitched_results)} pair(s) stitched")
        except Exception as e:
            logger.warning(f"Boundary stitching pass failed: {e}")

    # ── Phase 2: consolidate all chunk observations ────────────────────────────
    # Collect raw text and per-chunk data observations (sorted by sequence for ordered output)
    raw_text_parts: List[str] = []
    all_data_fragments: List[dict] = []

    for cr in sorted(chunk_results, key=lambda c: (c.get("page_number", 1), c.get("sequence", 0))):
        ext = cr.get("extracted") or {}
        raw = ext.get("raw_text", "") if isinstance(ext, dict) else ""
        if raw:
            raw_text_parts.append(f"[{cr.get('region', '')}, p{cr.get('page_number', 1)}, seq{cr.get('sequence', 0)}] {raw}")
        data = ext.get("data") if isinstance(ext, dict) else None
        if isinstance(data, dict):
            all_data_fragments.append(data)

    # Append stitched raw text after individual chunk text so consolidation sees it clearly
    for sr in stitched_results:
        if sr.get("stitched_raw"):
            ids_label = "-".join(str(cid) for cid in sr.get("chunk_ids", []))
            raw_text_parts.append(f"[STITCHED {ids_label}] {sr['stitched_raw']}")

    successful   = [cr for cr in chunk_results if cr.get("extracted")]

    # Merge data fragments:
    # - Scalar fields:  first non-null value wins
    # - Array fields:   concatenate all non-null arrays from every chunk
    # Stitched values then overwrite scalars (more complete — assembled from two images).
    merged_data: dict = {}
    array_accumulator: dict = {sid: [] for sid in _array_sub_steps}

    for frag in all_data_fragments:
        for k, v in frag.items():
            if k in _array_sub_steps:
                # Accumulate: parse string arrays too
                entries = v if isinstance(v, list) else None
                if entries is None and isinstance(v, str):
                    try:
                        parsed = json.loads(v)
                        entries = parsed if isinstance(parsed, list) else None
                    except Exception:
                        pass
                if entries:
                    array_accumulator[k].extend(entries)
            else:
                if v is not None and k not in merged_data:
                    merged_data[k] = v

    # Deduplicate accumulated arrays by (label, page) and store in merged_data
    for sid, entries in array_accumulator.items():
        if entries:
            seen_keys: set = set()
            deduped: list = []
            for entry in entries:
                if isinstance(entry, dict):
                    label = str(entry.get("label") or "").upper().strip()
                    sym   = str(entry.get("symbol_description") or "").strip()
                    page  = entry.get("page", 1)
                    # Only deduplicate when at least one identifying field is non-empty;
                    # entries with both label and symbol empty are kept as placeholders.
                    if label:
                        key = ("lbl", label, page)
                    elif sym:
                        key = ("sym", sym[:60], page)
                    else:
                        key = None  # always keep — can't distinguish
                    if key is None or key not in seen_keys:
                        if key is not None:
                            seen_keys.add(key)
                        deduped.append(entry)
                else:
                    deduped.append(entry)
            # Apply reference table: override symbol_description and set category for known labels
            for entry in deduped:
                if not isinstance(entry, dict):
                    continue
                lbl_upper = (entry.get("label") or "").strip().upper()
                # Try exact match, then check if label ENDS WITH a known key (handles "EXISTING 9.145m WIDTH ACME ENERGY EASEMENT")
                ref = _SYMBOL_REF.get(lbl_upper)
                if ref is None:
                    for ref_key, ref_val in _SYMBOL_REF.items():
                        if lbl_upper.endswith(ref_key.upper()):
                            ref = ref_val
                            break
                if ref:
                    entry["symbol_description"] = ref[0]
                    if not entry.get("category"):
                        entry["category"] = ref[1]
            merged_data[sid] = deduped

    for sr in stitched_results:
        for k, v in (sr.get("data") or {}).items():
            if v is not None:
                merged_data[k] = v  # stitched value wins over single-chunk value

    combined_raw = "\n\n".join(raw_text_parts)

    # ── Phase 3: final consolidation LLM pass (if sub_steps present) ──────────
    # Feed all raw observations back to the model to produce the final structured output.
    # Use head + tail budget so late chunks (where NOTES often appear) are never silently cut.
    if step.sub_steps and (combined_raw or merged_data):
        # Build required_keys_str: include output_format hints for structured sub-steps
        _rk_lines = []
        for line, sid in zip(_sub_step_schema_lines, _sub_step_ids):
            if sid in _p1_output_formats:
                _rk_lines.append(f'{line}\n    [OUTPUT FORMAT: {_p1_output_formats[sid][:400]}]')
            else:
                _rk_lines.append(line)
        required_keys_str = "\n".join(_rk_lines)
        keys_list = ", ".join(f'"{k}"' for k in _sub_step_ids)
        num_chunks    = len(chunks)
        num_stitched  = len(stitched_results)

        _MAX_RAW = 12000
        if len(combined_raw) > _MAX_RAW:
            _head = combined_raw[:8000]
            _tail = combined_raw[-3000:]
            _omitted = len(combined_raw) - 11000
            combined_raw_for_prompt = (
                _head +
                f"\n\n[... {_omitted} characters omitted from middle — "
                f"head and tail preserved to retain both early and late chunk content ...]\n\n" +
                _tail
            )
        else:
            combined_raw_for_prompt = combined_raw

        # Build notes-specific assembly hint for list-type sub-steps
        notes_hint = ""
        notes_ids = [sid for sid in _sub_step_ids if "note" in sid.lower()]
        if notes_ids:
            notes_hint = (
                f"\nNOTES ASSEMBLY — for fields: {', '.join(notes_ids)}\n"
                f"The NOTES section on engineering drawings is typically a numbered list "
                f"(e.g. 1., 2., 3. ...) that may span 3–6 adjacent image chunks.\n"
                f"Collect EVERY note number and its full text from ALL chunks in order.\n"
                f"Return the complete list as a single string with each note on its own line, "
                f"e.g.: \"1. All HV cable to comply...\\n2. Earthing to be designed...\\n3. New sub by ARP...\"\n"
                f"De-duplicate any note that appears in both a chunk and a [STITCHED ...] fragment.\n"
            )

        # Build legend-specific assembly hint for array-type sub-steps
        array_hint = ""
        array_ids = [sid for sid in _sub_step_ids if sid in _array_sub_steps]
        if array_ids:
            pre_merged_counts = {sid: len(merged_data.get(sid, [])) for sid in array_ids}
            # Build a summary of which entries are missing labels so the consolidation can fix them
            missing_label_summaries = {}
            for sid in array_ids:
                entries = merged_data.get(sid, [])
                missing = [e for e in entries if isinstance(e, dict) and not (e.get("label") or "").strip()]
                missing_label_summaries[sid] = len(missing)
            array_hint = (
                f"\nLEGEND/ARRAY ASSEMBLY — for fields: {', '.join(array_ids)}\n"
                f"The PARTIAL STRUCTURED DATA above already contains pre-merged arrays accumulated "
                f"from ALL {num_chunks} image chunks. Entry counts: "
                + ", ".join(f"{sid}={pre_merged_counts[sid]} entries ({missing_label_summaries[sid]} missing labels)" for sid in array_ids) +
                f"\nIMPORTANT: Many entries have symbol_description filled in but empty labels. "
                f"The ALL TEXT section above contains the actual label text for these symbols — "
                f"the legend rows appear in sequence in the raw text (e.g. 'NEW LV TRENCH', 'REMOVE LANTERN', etc.). "
                f"Match the symbol entries (in order) with the label text lines (in order) from the raw text "
                f"to complete each entry. The symbols and labels appear in the same sequential order in the legend. "
                f"OUTPUT the complete merged array with all symbol_description and label fields filled in. "
                f"De-duplicate by (label, page) when both are present. "
                f"CRITICAL: symbol_description must describe the VISUAL GEOMETRY of the drawn symbol "
                f"(e.g. 'Small hollow circle', 'Long dashed line', 'Hollow rectangle outline') — "
                f"NEVER a paraphrase of the label. Do NOT hallucinate entries.\n"
            )

        consolidation_prompt = (
            f"You are an expert engineering document analyst.\n\n"
            f"DOCUMENT: {fe.filename} ({page_size}, {strategy} chunking, {total_pages} page(s), "
            f"{num_chunks} chunks processed, {num_stitched} boundary pair(s) stitched)\n\n"
            f"ALL TEXT EXTRACTED FROM DOCUMENT CHUNKS (in sequence order):\n"
            f"{combined_raw_for_prompt}\n\n"
            f"PARTIAL STRUCTURED DATA FROM CHUNKS:\n{json.dumps(merged_data, indent=2)[:10000]}\n\n"
            f"TASK: {step.step_name}\n"
            f"{('DETAILS: ' + (step.details or '')[:300]) if step.details else ''}\n\n"
            f"CROSS-CHUNK ASSEMBLY — CRITICAL:\n"
            f"The text above was extracted from {num_chunks} separate image tiles of a large drawing.\n"
            f"- Lines marked with \u2192 were cut off at the right/bottom tile boundary "
            f"(content continues in the next tile).\n"
            f"- Lines marked with \u2190 continue from the previous tile.\n"
            f"- Sections labelled [STITCHED ...] are pre-assembled boundary fragments — "
            f"treat these as authoritative and prefer them over individual chunk fragments.\n"
            f"- For ALL list-type or multi-sentence fields: assemble the complete value by "
            f"reading ALL chunks in sequence order and joining partial entries at boundaries.\n"
            f"- Do NOT stop after the first chunk that mentions a field — read every chunk.\n"
            f"- De-duplicate entries that appear in both chunk text and stitched fragments.\n"
            f"{notes_hint}"
            f"{array_hint}\n"
            f"Using ALL the extracted text above, produce the final JSON extraction.\n"
            f"You MUST return a JSON object with EXACTLY these keys (no others):\n"
            f"{required_keys_str}\n\n"
            f"Rules:\n"
            f"- Search the full extracted text carefully for each field\n"
            f"- Each value should be the best extracted text or data for that field\n"
            f"- If a value is genuinely not present anywhere in the document, use null\n"
            f"- Do NOT wrap output in markdown fences — return raw JSON only\n"
            f"Required keys: {keys_list}"
        )
        try:
            consolidated = _llm_call(consolidation_prompt)
        except Exception as e:
            logger.warning(f"Consolidation LLM call failed: {e}. Using merged chunk data.")
            consolidated = merged_data
        # Post-consolidation protection for array sub-steps:
        # If the LLM dropped entries vs the pre-merged data, restore or inject missing entries.
        if isinstance(consolidated, dict):
            for sid in _array_sub_steps:
                pre = merged_data.get(sid)
                post = consolidated.get(sid)
                if not isinstance(pre, list):
                    continue
                if not isinstance(post, list) or len(post) < len(pre):
                    # LLM dropped too many — restore entirely
                    logger.info(
                        f"Array protection: restoring {sid} from {len(post) if isinstance(post, list) else '?'} "
                        f"→ {len(pre)} entries (LLM dropped entries)"
                    )
                    consolidated[sid] = pre
                else:
                    # Inject any pre-merged entries whose label is absent from LLM output
                    post_labels = {
                        (e.get("label") or "").strip().upper()
                        for e in post if isinstance(e, dict)
                    }
                    for entry in pre:
                        if not isinstance(entry, dict):
                            continue
                        lbl = (entry.get("label") or "").strip().upper()
                        if lbl and lbl not in post_labels:
                            post.append(entry)
                            post_labels.add(lbl)
                            logger.info(
                                f"Array injection: '{entry.get('label')}' missing from LLM output, restored"
                            )
    else:
        consolidated = merged_data

    # ── Phase 4: verification and gap-fill ────────────────────────────────────
    # Programmatically detect extraction gaps, then run a targeted LLM fill pass.
    consolidated = _verify_and_fill_extraction(
        consolidated, combined_raw, _sub_step_ids, _sub_step_schema_lines
    )

    # ── Phase 5: text-only analysis pass for Phase 2 sub-steps ───────────────
    # Phase 2 sub-steps are NOT extracted from images — they analyse the already-
    # consolidated Phase 1 text (drawing notes) via a plain text LLM call.
    # This ensures the analysis LLM sees the COMPLETE assembled notes rather than
    # partial chunks, and doesn't need to interpret images.
    p2_consolidated: dict = {}
    if _p2_ids:
        # Use the Phase 1 notes as the primary source — prefer the explicit notes
        # sub-step value if present, otherwise fall back to the full combined raw text
        notes_sid = next((sid for sid in _p1_ids if "note" in sid.lower()), None)
        p1_notes = (
            consolidated.get(notes_sid)
            if (notes_sid and isinstance(consolidated, dict) and consolidated.get(notes_sid))
            else combined_raw
        )

        p2_schema_str = "\n".join(_p2_schema)
        p2_keys_list  = ", ".join(f'"{k}"' for k in _p2_ids)

        # Build per-field instruction blocks — include output_format when defined
        p2_details_parts = []
        for ss in _p2_steps:
            sid    = ss.get("sub_step_id", "")
            sname  = ss.get("sub_step_name", sid)
            det    = ss.get("details", "")
            ofmt   = ss.get("output_format", "")
            req_tmpl = ss.get("required_note_template", "") or ss.get("required_note_text", "")
            block = f'Field "{sid}" — {sname}:\n{det}'
            if req_tmpl:
                block += f'\nRequired note text: "{req_tmpl}"'
            if ofmt:
                block += f'\nOutput format: {ofmt}'
            p2_details_parts.append(block)
        p2_details_str = "\n\n".join(p2_details_parts)

        # Budget combined_raw with head+tail so Phase 2 can find Substation Notes
        # annotation blocks that appear outside the numbered NOTES section
        _p2_MAX = 12000
        if len(combined_raw) > _p2_MAX:
            _p2_head = combined_raw[:8000]
            _p2_tail = combined_raw[-3000:]
            _p2_raw  = _p2_head + f"\n\n[...middle omitted...]\n\n" + _p2_tail
        else:
            _p2_raw = combined_raw

        p2_prompt = (
            f"You are an expert engineering document analyst.\n\n"
            f"DOCUMENT: {fe.filename}\n\n"
            f"ASSEMBLED NUMBERED NOTES AND CALLOUTS FROM THIS DRAWING:\n"
            f"{p1_notes}\n\n"
            f"FULL RAW TEXT FROM ALL DRAWING REGIONS (includes annotation blocks "
            f"outside the NOTES section, such as Substation Notes tables, funding "
            f"tables, and title block annotations):\n"
            f"{_p2_raw}\n\n"
            f"TASK: Analyse the text above and produce structured outputs for the "
            f"following fields. Read each field's instructions carefully.\n\n"
            f"{p2_details_str}\n\n"
            f"CRITICAL OUTPUT RULES:\n"
            f"- For 'sub-step-substation-data': search the FULL RAW TEXT for a "
            f"'Substation Notes', 'Sub Notes', or nearby annotation block listing "
            f"transformer size, switchgear, voltage, etc. Use EXACTLY these seven "
            f"field name labels in the output:\n"
            f"  Substation Asset Number, Transformer Size, HV Switchgear, "
            f"Voltage Level, LV Switchgear, Cubicle Size, Earthing\n"
            f"  Do not rename, add, or remove any field label.\n"
            f"- For 'sub-step-field-check': the required note is note 11 or similar. "
            f"The answer immediately follows the question as 'YES' or 'NO'. "
            f"If the note shows 'YES/NO' printed together (both options listed, "
            f"none crossed out or circled), the answer has NOT been filled in — "
            f"return exactly: 'NOT COMPLETED — note present but answer not filled in'. "
            f"If answered YES return 'YES'. If answered NO return "
            f"'NO — NON-COMPLIANT'. If note is absent return "
            f"'NOT FOUND — MISSING: this compliance note is mandatory'.\n"
            f"- For easement fields: return full verbatim note text with lot number(s) "
            f"if found; 'NOT REQUIRED' if not applicable; "
            f"'NOT FOUND — REQUIRED' if applicable but absent.\n"
            f"- Do NOT wrap output in markdown fences — return raw JSON only\n"
            f"Required keys: {p2_keys_list}"
        )

        try:
            p2_consolidated = _llm_call(p2_prompt)
            if not isinstance(p2_consolidated, dict):
                p2_consolidated = {}
            logger.info(f"Phase 2 analysis pass complete — {len(p2_consolidated)} field(s) produced")
        except Exception as e:
            logger.warning(f"Phase 2 analysis pass failed: {e}")
            p2_consolidated = {}

    # Merge Phase 1 and Phase 2 results into a single consolidated dict
    all_consolidated: dict = {}
    if isinstance(consolidated, dict):
        all_consolidated.update(consolidated)
    all_consolidated.update(p2_consolidated)

    # ── Apply reference table to array sub-step outputs ───────────────────────
    # Programmatically fix symbol_description and category for known legend labels,
    # regardless of what the consolidation LLM produced.
    for sid in _array_sub_steps:
        entries = all_consolidated.get(sid)
        if not isinstance(entries, list):
            continue
        # Also filter bad entries that slipped through (document references, etc.)
        _NON_LEGEND_SYM_RE = re.compile(
            r'cell\b.*\btable\b|in a table|at top of|horizontal line at top'
            r'|No symbol|no geometry|text label only|text only|label only'
            r'|Dash before text|dash.*text.*No geometry',
            re.IGNORECASE,
        )
        _NON_LEGEND_LBL_RE = re.compile(
            r"'\d{2}'|'23'|'24'|'25'"                      # construction ref numbers like '23'
            r'|\bFUNDED\b|\bCONTESTABLE\b'                 # cost table rows
            r'|\bCUSTOMER\b(?!\s+CONNECT)'                  # standalone CUSTOMER (not CONNECTION)
            r'|CO-ORDINATION SUPPLY|TO BE CONFIRMED'        # boilerplate
            r'|QUARTER OF 20\d\d|Third Quarter'             # dates
            r'|ASP LEVEL.*RETURN|RETURNED TO NEAREST'       # work instructions
            r'|CONDUCTOR\s+[A-Z]\s+\''                      # wiring notes like "CONDUCTOR B '23'"
            r'|\bLGA\b(?!\s+DEMARCATION)'                   # LGA area labels (not "LGA DEMARCATION")
            r'|DUCT USAGE CHARGES|USAGE CHARGES'            # cost table rows
            r'|(?<![A-Za-z])AVOUR ENERGY'                   # garbled "ACME ENERGY" (not "ACME ENERGY EASEMENT")
            r'|^NIL$|^-\s*NIL\b|^-\s+[A-Z]'               # "Nil", "- Nil", "- text" (list items)
            r'|PM SUB\s+\d+|ESTABLISHED ON ARP\d+'         # substation reference numbers
            r'|DESCRIPTION OF WORKS|SCOPE OF WORKS'         # notes section headers
            r'|DECOMMISSION EXISTING'                        # construction action verbs (not legend symbols)
            r'|POLE SUB\s+\d+'                              # pole substation references with numbers
            r'|\d+m/\d+kN'                                  # pole specifications (e.g. "14m/12kN")
            r'|\bAAC\b|\bABC\b',                            # conductor type codes (not legend entries)
            re.IGNORECASE,
        )
        cleaned_final: list = []
        seen_sym_page: set = set()   # deduplicate by (symbol_description, page) after ref-table lookup
        for entry in entries:
            if not isinstance(entry, dict):
                continue
            lbl = (entry.get("label") or "").strip()
            sym = (entry.get("symbol_description") or "").strip()
            # Drop entries with no label or single-char labels (e.g. north arrow "N")
            if not lbl or len(lbl) < 2:
                continue
            if _BAD_LABEL_RE.search(lbl) or len(lbl) > 70:
                continue
            # Drop entries that look like table rows or construction notes, not legend symbols
            if _NON_LEGEND_LBL_RE.search(lbl):
                continue
            if sym and _NON_LEGEND_SYM_RE.search(sym):
                continue
            lbl_upper = lbl.upper()
            ref = _SYMBOL_REF.get(lbl_upper)
            if ref is None:
                for ref_key, ref_val in _SYMBOL_REF.items():
                    if lbl_upper.endswith(ref_key.upper()):
                        ref = ref_val
                        break
            if ref:
                entry["symbol_description"] = ref[0]
                if not entry.get("category"):
                    entry["category"] = ref[1]
                # If matched via endswith (not exact key), normalize label to canonical form
                if _SYMBOL_REF.get(lbl_upper) is None:
                    for ref_key in _SYMBOL_REF:
                        if lbl_upper.endswith(ref_key.upper()):
                            entry["label"] = ref_key
                            break
            # Deduplicate by (symbol_description, page): keep first occurrence
            # (covers cases where two labels map to the same visual description)
            sym_after = (entry.get("symbol_description") or "").strip().upper()
            page_key  = entry.get("page", 1)
            sym_page_key = (sym_after, page_key)
            if sym_after and sym_page_key in seen_sym_page:
                continue
            if sym_after:
                seen_sym_page.add(sym_page_key)
            cleaned_final.append(entry)
        all_consolidated[sid] = cleaned_final

    # ── Assemble output in the same shape as _extract_file ────────────────────
    base = {
        "filename":              fe.filename,
        "filepath":              fe.filepath or "",
        "document_type":         fe.document_type,
        "document_category":     fe.document_category,
        "input_text_quality":    fe.text_quality,
        "document_quality":      "chunked-image",
        "page_size":             page_size,
        "requires_chunking":     True,
        "chunk_strategy":        strategy,
        "estimated_chunks":      len(chunks),
        "actual_chunks_processed": len(successful),
        "process_step_name":     step.step_name,
        "total_sections":        len(chunks),
        "relevant_sections":     len(successful),
        "raw_text_from_chunks":  combined_raw,
        "chunk_details":         chunk_results,
        "stitched_boundaries":   stitched_results,
    }

    if step.sub_steps:
        # Map each sub-step id to its value — Phase 1 from consolidated, Phase 2 from p2_consolidated
        sub_step_extractions = {}
        for ss in step.sub_steps:
            ss_id = ss.get("sub_step_id") or f"sub_step_{ss.get('sub_step_number', '')}"
            sub_step_extractions[ss_id] = {
                "value":          all_consolidated.get(ss_id),
                "sub_step_name":  ss.get("sub_step_name", ss_id),
                "phase":          ss.get("phase", 1),
            }
        base["sub_step_extractions"] = sub_step_extractions
    else:
        base["step_extraction"] = {"value": all_consolidated}

    return base


def _derive_tag_hint(sub_step_id: str) -> str:
    """Map a sub_step_id to a content_tag string used during section tagging."""
    tag_map = {
        "hv":          "hv",
        "lv":          "lv",
        "sl":          "sl",
        "earthing":    "earthing",
        "easement":    "easement",
        "substation":  "substation",
        "funding":     "funding",
        "transformer": "transformer",
        "voltage":     "voltage",
        "cubicle":     "cubicle",
        "switchgear":  "switchgear",
        "siteplan":    "siteplan",
    }
    sid_lower = sub_step_id.lower()
    for key, tag in tag_map.items():
        if key in sid_lower:
            return tag
    return ""


def _extract_per_sub_step(all_tagged: list, sub_steps: list) -> dict:
    """
    Run a separate targeted extraction for each sub-step.

    Per-sub-step behaviour:
    - Low-threshold categories (funding, easement) use score >= 0.25 so short
      checklist-style sections are not filtered out.
    - All other categories use score >= 0.4.
    - Sections are narrowed by content_tag match when possible.
    - When no tag match exists, fall back to the top 5 by relevance score only
      (not all sections) to prevent unrelated content flooding the prompt.
    - Checklist categories (easement, earthing) also force-include the first two
      document sections regardless of score, since these items often appear as
      short header lines near the top of the document.
    """
    # Categories with lower relevance threshold (short/sparse but critical content)
    LOW_THRESHOLD_TAGS = {"funding", "easement"}
    # Categories that additionally force-include the first two document sections
    CHECKLIST_TAGS = {"easement", "earthing"}

    # Build (ss_id, ss_relevant, ss_step) tuples for all sub-steps up front,
    # then run all structured extractions in parallel.
    tasks: List[tuple] = []
    for ss in sub_steps:
        ss_id      = ss.get("sub_step_id") or f"sub_step_{ss.get('sub_step_number', '')}"
        ss_name    = ss.get("sub_step_name", ss_id)
        ss_details = ss.get("details", "")
        ss_inst    = ss.get("instructional_sub_steps") or []

        ss_step = ProcessStepDef(
            step_id=ss_id,
            step_name=ss_name,
            details=ss_details,
            instructional_sub_steps=ss_inst if ss_inst else None,
        )

        tag_hint  = _derive_tag_hint(ss_id)
        threshold = 0.25 if tag_hint in LOW_THRESHOLD_TAGS else 0.4

        above_threshold = [s for s in all_tagged if s.get("relevance_score", 0) >= threshold]

        if tag_hint:
            tag_filtered = [s for s in above_threshold if tag_hint in s.get("content_tags", [])]
            if tag_filtered:
                ss_relevant = tag_filtered
            else:
                ss_relevant = sorted(
                    above_threshold,
                    key=lambda s: s.get("relevance_score", 0),
                    reverse=True,
                )[:5]
        else:
            ss_relevant = above_threshold

        if tag_hint in CHECKLIST_TAGS:
            first_two = all_tagged[:2]
            existing_names = {s.get("section_name") for s in ss_relevant}
            for s in first_two:
                if s.get("section_name") not in existing_names:
                    ss_relevant = [s] + ss_relevant

        tasks.append((ss_id, ss_relevant, ss_step))

    # Run all sub-step extractions concurrently — each is an independent LLM call.
    results = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(tasks)) as executor:
        future_to_id = {
            executor.submit(_structured_extraction, ss_rel, ss_stp): ss_id
            for ss_id, ss_rel, ss_stp in tasks
        }
        for future in concurrent.futures.as_completed(future_to_id):
            sid = future_to_id[future]
            try:
                results[sid] = future.result()
            except Exception as e:
                logger.error(f"Sub-step extraction failed for {sid}: {e}")
                results[sid] = {"error": str(e)}

    return results


# ── PDF extraction ────────────────────────────────────────────────────────────

def _extract_pdf_sections(filepath: str, chunk_page_size: Optional[int] = None):
    """
    Split a PDF into logical sections using heading detection.

    Returns (sections_list, quality_flag) where quality_flag is one of:
      'text'    — normal text PDF, section-level split used
      'sparse'  — low text density, fell back to page-level grouping
      'drawing' — very low alpha ratio (likely CAD/drawing), minimal text
      'chunked' — forced page-level split due to large page size

    Each section dict: {section_number, section_name, page_number, text, tables, images}
    chunk_page_size: when set, forces page-level grouping with this group size (e.g. 1 for large drawings).
    """
    try:
        with pdfplumber.open(filepath) as pdf:
            # ── Pass 1: collect all page text and tables ──────────────────────
            page_data = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                tables = []
                for tbl in (page.extract_tables() or []):
                    if tbl and len(tbl) > 1:
                        tables.append({
                            "headers": [str(h or "") for h in tbl[0]],
                            "rows":    [[str(c or "") for c in row] for row in tbl[1:]],
                        })
                page_data.append({"page": page.page_number, "text": text, "tables": tables})

    except Exception as e:
        return ([{
            "section_number": None, "section_name": "Extraction Error",
            "page_number": 0, "text": str(e), "tables": [], "images": [],
        }], "error")

    # ── Force chunk-level split for large pages ───────────────────────────────
    if chunk_page_size is not None:
        sections = _page_level_sections(page_data, group_size=chunk_page_size)
        return sections, "chunked"

    # ── Assess overall document text quality ─────────────────────────────────
    all_text = "\n".join(p["text"] for p in page_data)
    overall_alpha = _alpha_ratio(all_text)

    # Drawing / CAD PDF: very low alpha ratio — return page-level sections
    if overall_alpha < 0.25:
        sections = _page_level_sections(page_data, group_size=2)
        return sections, "drawing"

    # ── Pass 2: heading-based section split ───────────────────────────────────
    sections: list = []
    current = _new_section(None, "Preamble", 1)

    for pd in page_data:
        pnum = pd["page"]
        for line in pd["text"].split("\n"):
            mn = _NUMBERED.match(line)
            mf = _FORCED_HEADINGS.match(line) if not mn else None
            mc = _ALLCAPS.match(line) if not mn and not mf else None

            if mn:
                _flush(sections, current)
                current = _new_section(mn.group(1), mn.group(2).strip(), pnum)
            elif mf:
                # High-value keyword heading — always split regardless of capitalisation
                _flush(sections, current)
                current = _new_section(None, line.strip(), pnum)
            elif mc and _is_valid_heading(mc.group(1)):
                _flush(sections, current)
                current = _new_section(None, mc.group(1).strip(), pnum)
            else:
                current["text"] += line + "\n"

        for tbl in pd["tables"]:
            current["tables"].append(tbl)

    _flush(sections, current)

    # ── Fallback: too many sections → page-level grouping ────────────────────
    if len(sections) > MAX_SECTIONS_BEFORE_FALLBACK:
        sections = _page_level_sections(page_data, group_size=3)
        return sections, "sparse"

    if not sections:
        sections = [{
            "section_number": None, "section_name": "Full Document",
            "page_number": 1, "text": all_text[:8000],
            "tables": [], "images": [],
        }]

    return sections, "text"


def _page_level_sections(page_data: list, group_size: int = 3) -> list:
    """
    Group consecutive pages into sections as a fallback for documents
    where heading detection produces too many or too few sections.
    """
    sections = []
    for i in range(0, len(page_data), group_size):
        chunk = page_data[i:i + group_size]
        combined_text = "\n\n".join(p["text"] for p in chunk if p["text"].strip())
        tables = [t for p in chunk for t in p["tables"]]
        if combined_text.strip() or tables:
            pstart = chunk[0]["page"]
            pend   = chunk[-1]["page"]
            name   = f"Pages {pstart}–{pend}" if pstart != pend else f"Page {pstart}"
            sections.append({
                "section_number": None, "section_name": name,
                "page_number": pstart, "text": combined_text,
                "tables": tables, "images": [],
            })
    return sections


def _new_section(num, name, page):
    return {"section_number": num, "section_name": name, "page_number": page,
            "text": "", "tables": [], "images": []}


def _flush(sections, current):
    text = current["text"].strip()
    # Only flush if the text has meaningful alphabetic content
    if (text and _alpha_ratio(text) >= MIN_ALPHA_RATIO) or current["tables"]:
        sections.append(dict(current))


# ── Spreadsheet extraction ────────────────────────────────────────────────────

_SPREADSHEET_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.xlsb'}


def _read_spreadsheet_as_tables(filepath: str) -> list:
    """
    Read an Excel workbook and return a list of worksheet dicts:
        {"sheet_name": str, "headers": [...], "rows": [[...], ...]}
    Supports .xlsx/.xlsm/.xlsb via openpyxl and .xls via xlrd.
    """
    ext = Path(filepath).suffix.lower()
    sheets = []
    try:
        if ext in {'.xlsx', '.xlsm', '.xlsb'}:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = []
                for row in ws.iter_rows(values_only=True):
                    all_rows.append([str(c) if c is not None else "" for c in row])
                # Drop fully-empty rows
                all_rows = [r for r in all_rows if any(c.strip() for c in r)]
                if not all_rows:
                    continue
                headers = all_rows[0]
                data_rows = all_rows[1:]
                sheets.append({"sheet_name": sheet_name, "headers": headers, "rows": data_rows})
            wb.close()
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for idx in range(wb.nsheets):
                ws  = wb.sheet_by_index(idx)
                name = wb.sheet_names()[idx]
                all_rows = []
                for i in range(ws.nrows):
                    all_rows.append([str(ws.cell_value(i, j)) for j in range(ws.ncols)])
                all_rows = [r for r in all_rows if any(c.strip() for c in r)]
                if not all_rows:
                    continue
                headers = all_rows[0]
                data_rows = all_rows[1:]
                sheets.append({"sheet_name": name, "headers": headers, "rows": data_rows})
    except Exception as e:
        logger.error(f"_read_spreadsheet_as_tables failed for {filepath}: {e}")
    return sheets


def _select_target_worksheets(sheets: list, filename: str) -> list:
    """
    Pass 1: Ask Claude to review all worksheet names + headers and return
    the list of sheet names that contain asset record data worth extracting.
    Returns a list of sheet names to process.
    """
    sheet_summaries = []
    for s in sheets:
        header_preview = "\t".join(s["headers"][:20])  # first 20 columns
        row_count      = len(s["rows"])
        # Show first 3 data rows as a sample
        sample_rows = "\n".join("\t".join(r[:20]) for r in s["rows"][:3])
        sheet_summaries.append(
            f"Sheet: {s['sheet_name']} ({row_count} data rows)\n"
            f"Headers: {header_preview}\n"
            f"Sample rows:\n{sample_rows}"
        )

    prompt = (
        f"You are reviewing a workbook named '{filename}' to identify which worksheets "
        f"contain asset register data suitable for extraction.\n\n"
        f"WORKSHEETS IN THIS FILE:\n\n"
        + "\n\n---\n\n".join(sheet_summaries)
        + "\n\n"
        "Asset register worksheets typically contain rows of individual assets with columns for "
        "identifiers (Asset ID, Tag, FLOC), asset type or description, location, status, "
        "technical specifications (voltage, kVA, capacity), manufacturer/model, condition, "
        "or installation/maintenance dates.\n\n"
        "Exclude worksheets that are: instructions/cover sheets, lookup tables, charts-only, "
        "summary/pivot tables, or empty/placeholder tabs.\n\n"
        "Return ONLY a JSON object: "
        "{\"target_sheets\": [\"SheetName1\", \"SheetName2\"], "
        "\"reasoning\": \"one sentence explaining the selection\"}"
    )

    try:
        result = _llm_call(prompt)
        targets = result.get("target_sheets") or []
        reasoning = result.get("reasoning", "")
        logger.info(f"Worksheet selection for '{filename}': {targets} — {reasoning}")
        # Validate: only return names that actually exist in the workbook
        valid_names = {s["sheet_name"] for s in sheets}
        return [t for t in targets if t in valid_names]
    except Exception as e:
        logger.warning(f"Worksheet selection failed: {e}. Processing all sheets.")
        return [s["sheet_name"] for s in sheets]


def _extract_asset_spreadsheet(fe: FileEntry, filepath: str) -> dict:
    """
    Dedicated extraction for proc-extract-asset-spreadsheet.
    Pass 1 — Claude selects which worksheets contain asset data.
    Pass 2 — Claude extracts structured asset records from each targeted sheet.
    Returns a result dict with step_extraction.asset_records.
    """
    logger.info(f"_extract_asset_spreadsheet: filepath={filepath!r}, exists={Path(filepath).exists()}")

    # If the explicit filepath doesn't exist, try to locate the file by name
    if not Path(filepath).exists():
        # Search common container mount points
        search_roots = [Path("/documents"), Path("/app/documents"), Path("/data")]
        found = None
        for root in search_roots:
            if root.exists():
                matches = list(root.rglob(fe.filename))
                if matches:
                    found = str(matches[0])
                    logger.info(f"Found spreadsheet at: {found}")
                    break
        if found:
            filepath = found
        else:
            logger.error(f"Spreadsheet not found: {filepath!r} (also searched /documents, /app/documents, /data)")
            return {
                "filename": fe.filename, "filepath": filepath,
                "document_type": fe.document_type, "document_category": fe.document_category,
                "document_quality": "spreadsheet",
                "step_extraction": {"asset_records": [], "total_assets": 0,
                                    "sheets_processed": [],
                                    "note": f"File not found: {filepath}"},
            }

    sheets = _read_spreadsheet_as_tables(filepath)
    if not sheets:
        return {
            "filename": fe.filename, "filepath": filepath,
            "document_type": fe.document_type, "document_category": fe.document_category,
            "document_quality": "spreadsheet",
            "step_extraction": {"asset_records": [], "total_assets": 0,
                                "sheets_processed": [], "note": "No readable worksheets found."},
        }

    # Pass 1: let Claude decide which worksheets to target
    target_names   = _select_target_worksheets(sheets, fe.filename)
    target_sheets  = [s for s in sheets if s["sheet_name"] in target_names]
    sheets_skipped = [s["sheet_name"] for s in sheets if s["sheet_name"] not in target_names]
    if sheets_skipped:
        logger.info(f"Skipping non-asset worksheets: {sheets_skipped}")

    # Fallback: if selection returned nothing, process all sheets
    if not target_sheets:
        target_sheets = sheets

    all_asset_records: list = []
    sheets_processed: list  = []

    # Pass 2: extract asset records from each targeted sheet
    for sheet in target_sheets:
        data_rows = sheet["rows"]
        if not data_rows:
            continue

        sheets_processed.append(sheet["sheet_name"])

        # Process rows in batches to stay within Claude's output token limit
        BATCH_SIZE  = 50
        header_line = "\t".join(sheet["headers"])
        sheet_records: list = []
        batches = [data_rows[i:i + BATCH_SIZE] for i in range(0, min(len(data_rows), 500), BATCH_SIZE)]

        for batch_idx, batch in enumerate(batches):
            row_lines  = ["\t".join(r) for r in batch]
            table_text = header_line + "\n" + "\n".join(row_lines)

            prompt = (
                f"You are extracting structured asset records from a spreadsheet worksheet.\n\n"
                f"Worksheet: {sheet['sheet_name']}\n"
                f"Source file: {fe.filename}\n"
                f"Batch: rows {batch_idx * BATCH_SIZE + 1}–{batch_idx * BATCH_SIZE + len(batch)} "
                f"of {len(data_rows)}\n\n"
                f"SPREADSHEET DATA (tab-separated, first row is header):\n"
                f"{table_text}\n\n"
                "For EVERY data row, extract a JSON asset record with these fields "
                "(use null when not present in the spreadsheet):\n"
                "  asset_id, serial_number, asset_type, description, location, address,\n"
                "  feeder, voltage_level, capacity_rating, manufacturer, model,\n"
                "  condition_rating, status, installation_date, last_inspection_date\n\n"
                "CRITICAL — asset_id mapping rule:\n"
                "  Use the 'Superior Functional Location' column as asset_id. "
                "If that column is absent or empty for a row, fall back to any column "
                "named 'Functional Location', 'FLOC', 'Asset ID', or 'Tag'.\n\n"
                "Map all other spreadsheet columns to the remaining fields as best you can. "
                "Do NOT skip rows — include one record per data row even if some fields are null.\n\n"
                "Return ONLY a JSON object: "
                "{\"asset_records\": [{...}, ...], \"total_assets\": <int>}"
            )

            try:
                result  = _llm_call(prompt)
                records = result.get("asset_records") or []
                for rec in records:
                    rec["source_sheet"]    = sheet["sheet_name"]
                    rec["source_document"] = fe.filename
                sheet_records.extend(records)
            except Exception as e:
                logger.error(f"Asset extraction failed for sheet '{sheet['sheet_name']}' batch {batch_idx}: {e}")

        all_asset_records.extend(sheet_records)
        logger.info(f"Sheet '{sheet['sheet_name']}': {len(sheet_records)} asset records extracted "
                    f"({len(batches)} batch(es)).")

    return {
        "filename":          fe.filename,
        "filepath":          filepath,
        "document_type":     fe.document_type,
        "document_category": fe.document_category,
        "document_quality":  "spreadsheet",
        "sheets_processed":  sheets_processed,
        "sheets_skipped":    sheets_skipped if sheets_skipped else [],
        "step_extraction": {
            "asset_records":    all_asset_records,
            "total_assets":     len(all_asset_records),
            "sheets_processed": sheets_processed,
        },
    }


# ── Plain-text extraction ─────────────────────────────────────────────────────

def _extract_text_file(filepath: str) -> list:
    try:
        with open(filepath, "r", errors="replace") as f:
            content = f.read()
        return [{
            "section_number": None,
            "section_name":   "Full Document",
            "page_number":    1,
            "text":           content[:10000],
            "tables":         [],
            "images":         [],
        }]
    except Exception as e:
        return [{
            "section_number": None,
            "section_name":   "Error",
            "page_number":    0,
            "text":           str(e),
            "tables":         [],
            "images":         [],
        }]


# ── Section relevance tagging (chunked batches) ───────────────────────────────

def _tag_all_sections_chunked(sections: list, step: ProcessStepDef) -> list:
    """
    Tag sections in batches of TAGGING_BATCH_SIZE to avoid context limit failures.
    Empty and low-quality sections are skipped and given score 0.0.
    """
    tagged = list(sections)

    # Classify each section as taggable or skip.
    # Sections with tables are always taggable regardless of alpha ratio —
    # technical drawings have low alpha (codes, numbers) but rich table content.
    taggable_indices = []
    for idx, s in enumerate(sections):
        text = (s.get("text") or "").strip()
        has_tables = bool(s.get("tables"))
        if (text and _alpha_ratio(text) >= MIN_ALPHA_RATIO) or has_tables:
            taggable_indices.append(idx)
        else:
            tagged[idx] = {**sections[idx],
                           "relevance_score": 0.0,
                           "relevance_reason": "Low text quality or empty",
                           "content_tags": []}

    if not taggable_indices:
        return tagged

    # Process in batches
    for batch_start in range(0, len(taggable_indices), TAGGING_BATCH_SIZE):
        batch_indices = taggable_indices[batch_start:batch_start + TAGGING_BATCH_SIZE]
        batch_tags = _tag_batch(sections, batch_indices, step)
        for rank, idx in enumerate(batch_indices):
            tags = batch_tags[rank] if rank < len(batch_tags) else {}
            tagged[idx] = {**sections[idx],
                           "relevance_score":  tags.get("relevance_score", 0.0),
                           "relevance_reason": tags.get("relevance_reason", "Tagging failed"),
                           "content_tags":     tags.get("content_tags", [])}

    return tagged


def _tag_batch(sections: list, indices: list, step: ProcessStepDef) -> list:
    """Tag a single batch of sections with one LLM call."""
    section_blocks = []
    for rank, idx in enumerate(indices):
        s = sections[idx]
        # Use up to 600 chars of text preview
        preview = (s.get("text") or "")[:600].strip()
        # Append table headers so the LLM can see structured content in drawings
        # (technical drawings have low text but rich table data)
        table_preview = ""
        for tbl in (s.get("tables") or [])[:4]:
            headers = " | ".join(str(h) for h in (tbl.get("headers") or []) if h)
            if headers:
                table_preview += f"\n[Table: {headers}]"
        section_blocks.append(
            f"[{rank}] Section: {s.get('section_name')} (page {s.get('page_number')})\n"
            f"{preview}{table_preview}"
        )

    prompt = (
        f"You are tagging document sections for relevance to a process step.\n\n"
        f"Process step: {step.step_name}\n"
        f"Step details: {(step.details or step.summary or '').strip()[:500]}\n\n"
        f"Tag each of the {len(indices)} sections below.\n\n"
        + "\n\n---\n\n".join(section_blocks)
        + "\n\n"
        "Respond with a JSON object: {\"sections\": [<one entry per section in order>]}\n"
        "Each entry: {\"relevance_score\": 0.0-1.0, \"relevance_reason\": \"one sentence\", "
        "\"content_tags\": [\"short\", \"identifiers\"]}\n"
        "Tags: short lowercase identifiers for topics that genuinely appear in the section "
        "(e.g. 'hv', 'lv', 'sl', 'earthing', 'easement', 'substation', "
        "'funding', 'scope', 'requirements', 'schedule', 'cost', "
        "'transformer', 'voltage', 'cubicle', 'switchgear', 'siteplan'). "
        "Only include tags where the content actually contains that topic.\n"
        "Important tagging rules:\n"
        "- Tag 'funding' for any section mentioning: determination of funding, capital contribution, "
        "contestable works, non-contestable works, ancillary network services fees, reimbursement, "
        "administration fee, design fee, connection offer fee — even if the section is short.\n"
        "- Tag 'easement' for any section mentioning: land interests, easement, LIG, "
        "land interest guidelines — even if the value is 'NA' or 'nil'.\n"
        "- Tag 'siteplan' for any section containing site plan drawings, substation notes tables, "
        "electrical network diagrams, or asset annotations on a drawing.\n"
        "- Tag 'transformer' for any section mentioning transformer size, kVA, MVA ratings.\n"
        "- Tag 'switchgear' for any section mentioning HV switchgear, LV switchgear, RMU, cubicle, "
        "Siemens RLR, or switchboard specifications.\n"
        "- Tag 'voltage' for any section mentioning operating voltage level (kV).\n"
        "- Tag 'cubicle' for any section mentioning cubicle size, dimensions, or enclosure specs.\n"
        "- Assign relevance_score >= 0.5 to any section whose heading or content matches "
        "one of the above keywords, regardless of section length."
    )

    expected = len(indices)
    for attempt in range(LLM_MAX_RETRIES):
        try:
            result = _llm_call_fast(prompt)
            sections_out = result.get("sections", [])
            if len(sections_out) == expected:
                return sections_out
            # Count mismatch — pad or trim and log
            logger.warning(
                f"_tag_batch: expected {expected} section tags, got {len(sections_out)} "
                f"(attempt {attempt + 1}/{LLM_MAX_RETRIES})"
            )
            if attempt < LLM_MAX_RETRIES - 1:
                time.sleep(LLM_RETRY_DELAYS[attempt])
                continue
            # Last attempt: return what we have (pad with empty dicts if short)
            while len(sections_out) < expected:
                sections_out.append({})
            return sections_out[:expected]
        except Exception as e:
            logger.error(f"_tag_batch failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}")
            if attempt < LLM_MAX_RETRIES - 1:
                time.sleep(LLM_RETRY_DELAYS[attempt])
    return [{} for _ in indices]


# ── Structured step extraction ────────────────────────────────────────────────

def _structured_extraction(relevant_sections: list, step: ProcessStepDef) -> dict:
    """
    Given the relevant sections, ask the LLM to produce a structured extraction
    matching the expected output schema of the process step.

    Uses a two-pass approach for large documents:
      Pass 1 — extract from the highest-scored sections (up to ~8000 chars)
      Pass 2 — merge any additional lower-scored sections if context allows
    """
    if not relevant_sections:
        return {"note": "No relevant sections found for this process step."}

    # Sort by relevance score descending so highest-value content goes first
    sorted_sections = sorted(relevant_sections, key=lambda s: s.get("relevance_score", 0), reverse=True)

    # Build section blocks — use more text per section (2000 chars) for better fidelity
    parts = []
    total_chars = 0
    CHAR_BUDGET = 12000  # allow more context for the extraction prompt

    for s in sorted_sections:
        tags  = s.get("content_tags", [])
        text  = (s.get("text") or "")[:2000]
        tables_detail = ""
        if s.get("tables"):
            for tbl in s["tables"][:3]:  # include up to 3 tables per section
                headers = " | ".join(tbl.get("headers", []))
                rows    = "\n".join(" | ".join(r) for r in tbl.get("rows", [])[:10])
                tables_detail += f"\nTable:\n{headers}\n{rows}"
        block = (
            f"[Section: {s.get('section_name')} | Page: {s.get('page_number')} | "
            f"Score: {s.get('relevance_score', 0):.2f} | Tags: {tags}]\n"
            f"{text}{tables_detail}"
        )
        if total_chars + len(block) > CHAR_BUDGET:
            break
        parts.append(block)
        total_chars += len(block)

    sections_block = "\n\n---\n\n".join(parts)

    # Build a focused expected output description
    expected_desc = _build_expected_desc(step)

    # For sub-step extractions, add strict cross-category exclusion instructions
    # to prevent items from one voltage class or infrastructure type bleeding into another.
    scope_instruction = ""
    if step.step_id and step.step_id.startswith("sub-step-extract-"):
        tag_hint = _derive_tag_hint(step.step_id)
        if tag_hint:
            all_categories = ["hv", "lv", "sl", "earthing", "easement", "substation", "funding"]
            other_cats = [t.upper() for t in all_categories if t != tag_hint]
            scope_instruction = (
                f"- SCOPE: Extract ONLY items that explicitly relate to {tag_hint.upper()}. "
                f"Do NOT include items belonging to other categories "
                f"({', '.join(other_cats)}) even if they appear in the same section.\n"
                f"- If an item references a different voltage class or infrastructure type "
                f"(e.g. an HV cable in an LV section), exclude it.\n"
                f"- General design notes or 'method of supply' summaries are NOT {tag_hint.upper()} "
                f"requirements unless they explicitly describe {tag_hint.upper()} infrastructure.\n"
            )

    # Build instructional guidance from instructional_sub_steps when present
    inst_guidance = ""
    if step.instructional_sub_steps:
        inst_parts = []
        for k, ist in enumerate(step.instructional_sub_steps, 1):
            title = ist.get("title", "")
            instructions = (ist.get("instructions") or "").strip()
            if title or instructions:
                inst_parts.append(f"  {k}. {title}:\n     {instructions}")
        if inst_parts:
            inst_guidance = (
                "Detailed extraction instructions (follow these carefully):\n"
                + "\n".join(inst_parts)
                + "\n\n"
            )

    prompt = (
        f"You are extracting structured engineering data from document sections.\n\n"
        f"Task: {step.step_name}\n"
        f"Context: {(step.details or step.summary or '').strip()[:800]}\n\n"
        + inst_guidance
        + (f"Required output structure:\n{expected_desc}\n\n" if expected_desc else "")
        + f"Document sections (highest relevance first):\n\n{sections_block}\n\n"
        "Instructions:\n"
        "- Extract ALL items that match the required output categories\n"
        + scope_instruction
        + "- For each item include: 'description' (verbatim or close paraphrase from source), "
        "'source_section' (section name), 'source_page' (page number)\n"
        "- Include specific values: voltages (kV), quantities, distances (m), costs ($), "
        "reference numbers, asset IDs\n"
        "- Use the exact category keys from the required output structure\n"
        "- Do NOT invent data not present in the source text\n"
        "- Do NOT include metadata from prior workflow steps (no 'documents' list, no 'processing_plan')\n"
        "Return a JSON object with only the extraction categories."
    )

    try:
        return _llm_call(prompt)
    except Exception as e:
        logger.error(f"_structured_extraction failed after all retries: {e}")
        return {"error": str(e)}


def _build_expected_desc(step: ProcessStepDef) -> str:
    """Build a concise expected output description from the step definition."""
    if not step.expected_output:
        return ""
    fields = step.expected_output.get("fields")
    description = step.expected_output.get("description", "")
    # Only use fields dict if it looks like extraction categories (not pipeline schemas)
    pipeline_keys = {"documents", "scan_results", "processing_plan", "total_documents",
                     "matched_files", "total_files_scanned", "output_file"}
    if isinstance(fields, dict):
        clean_fields = {k: v for k, v in fields.items() if k not in pipeline_keys}
        if clean_fields:
            return json.dumps(clean_fields, indent=2)
    if description:
        # Strip pipeline-step references from the description
        lines = [l for l in description.split("|")
                 if not any(kw in l.lower() for kw in ["step 1", "step 2", "documentreview", "processingplan"])]
        return " | ".join(lines).strip()
    return ""


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8090)
