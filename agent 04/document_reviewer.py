import json
import logging
import os
import re
import sys
import time
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import openai

# ── Load tool: list_folder_files ───────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
from tools import load_tool as _load_tool
_list_files_tool = _load_tool("tool-list-folder-files")

load_dotenv()

logger = logging.getLogger("document_reviewer")
logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Document Reviewer Agent")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

MODEL = os.environ.get("LLM_MODEL", "gpt-4o")
_openai_client = openai.OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

# ── Robust LLM call with retry ─────────────────────────────────────────────────
_JSON_FENCE    = re.compile(r"```(?:json)?\s*([\s\S]*?)```", re.IGNORECASE)
_LLM_RETRIES   = 3
_LLM_DELAYS    = [2, 5, 10]

def _llm_call(prompt: str) -> Any:
    last_exc = None
    for attempt in range(_LLM_RETRIES):
        try:
            resp = _openai_client.chat.completions.create(
                model=MODEL,
                messages=[{"role": "user", "content": prompt}],
                response_format={"type": "json_object"},
                temperature=0,
            )
            text = resp.choices[0].message.content.strip()
            m = _JSON_FENCE.search(text)
            if m:
                text = m.group(1).strip()
            return json.loads(text)
        except Exception as e:
            last_exc = e
            delay = _LLM_DELAYS[min(attempt, len(_LLM_DELAYS) - 1)]
            logger.warning(f"LLM call failed (attempt {attempt+1}/{_LLM_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc

SUPPORTED_EXTENSIONS = {'.txt', '.pdf', '.docx', '.doc', '.csv', '.json', '.md', '.xlsx', '.xls', '.xlsm', '.xlsb'}


# ── TOOL: Read file content ────────────────────────────────────────────────────
def read_file_content(filepath: str, max_chars: int = 4000) -> str:
    """Read and extract text from a file. Supports PDF, DOCX, and plain text."""
    path = Path(filepath)
    ext = path.suffix.lower()
    try:
        if ext in {'.txt', '.md', '.csv', '.json'}:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()[:max_chars]

        elif ext == '.pdf':
            import pdfplumber
            text = ""
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages[:6]:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text[:max_chars] if text.strip() else "[PDF contained no extractable text]"

        elif ext in {'.docx', '.doc'}:
            from docx import Document
            doc = Document(filepath)
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return text[:max_chars]

        elif ext in {'.xlsx', '.xlsm', '.xlsb'}:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(max_row=50, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        rows.append("\t".join(cells))
                if rows:
                    parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            wb.close()
            return "\n\n".join(parts)[:max_chars] if parts else "[Spreadsheet contained no data]"

        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(filepath)
            parts = []
            for sheet in wb.sheets():
                rows = []
                for rx in range(min(50, sheet.nrows)):
                    cells = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
                    if any(c.strip() for c in cells):
                        rows.append("\t".join(cells))
                if rows:
                    parts.append(f"[Sheet: {sheet.name}]\n" + "\n".join(rows))
            return "\n\n".join(parts)[:max_chars] if parts else "[Spreadsheet contained no data]"

        else:
            return f"[Unsupported file type: {ext}]"

    except Exception as e:
        return f"[Error reading file: {str(e)}]"


# ── TOOL: Assess content quality ──────────────────────────────────────────────
def assess_content_quality(content: str, ext: str) -> dict:
    """
    Determine whether a file's extracted content is text or image-based,
    and recommend the best tool to read and extract data from it.

    Returns:
        content_type:       "text" | "image" | "unsupported" | "error"
        text_quality:       "high" | "low" | "none"
        recommended_reader: "direct" | "pdfplumber" | "python-docx" |
                            "ocr-required" | "unsupported" | "error"
        quality_note:       human-readable explanation
    """
    ext = ext.lower()

    # Non-extractable sentinel strings from read_file_content()
    if content.startswith("[PDF contained no extractable text]"):
        return {
            "content_type": "image",
            "text_quality": "none",
            "recommended_reader": "ocr-required",
            "quality_note": "PDF appears to be a scanned image. OCR is required to extract text."
        }
    if content.startswith("[Unsupported file type"):
        return {
            "content_type": "unsupported",
            "text_quality": "none",
            "recommended_reader": "unsupported",
            "quality_note": f"File type {ext} is not supported for text extraction."
        }
    if content.startswith("[Error reading file"):
        return {
            "content_type": "error",
            "text_quality": "none",
            "recommended_reader": "error",
            "quality_note": content
        }

    # Determine reader based on extension
    if ext in {'.txt', '.md', '.csv', '.json'}:
        reader = "direct"
    elif ext == '.pdf':
        reader = "pdfplumber"
    elif ext in {'.docx', '.doc'}:
        reader = "python-docx"
    elif ext in {'.xlsx', '.xlsm', '.xlsb'}:
        reader = "openpyxl"
    elif ext == '.xls':
        reader = "xlrd"
    else:
        reader = "direct"

    # Text quality based on character count
    char_count = len(content.strip())
    if char_count >= 500:
        quality = "high"
        note = f"Text-based document. {char_count} characters extracted. Reader: {reader}."
    elif char_count > 0:
        quality = "low"
        note = f"Limited text extracted ({char_count} characters). Document may be partially image-based. Reader: {reader}."
    else:
        quality = "none"
        note = "No text content extracted despite successful read attempt."

    return {
        "content_type": "text",
        "text_quality": quality,
        "recommended_reader": reader,
        "quality_note": note
    }


# ── TOOL: Check filename match ─────────────────────────────────────────────────
def check_filename_match(filename: str, search_context: str) -> dict:
    """Use LLM to assess whether a filename suggests it matches the search context."""
    prompt = (
        f"You are a document classification assistant.\n\n"
        f"Search context — the user is looking for:\n{search_context}\n\n"
        f"Filename to assess: {filename}\n\n"
        f"Analyse the filename ONLY (do not assume file contents). Determine:\n"
        f"1. Does the filename suggest this could be one of the documents being searched for?\n"
        f"2. If yes, what document type does it appear to be?\n"
        f"3. Confidence score (0.0 = definitely not, 1.0 = definitely yes)\n\n"
        f'Respond only in JSON: {{"match": true/false, "document_type": "...", "confidence": 0.0, "reasoning": "..."}}'
    )
    try:
        return _llm_call(prompt)
    except Exception as e:
        return {"match": False, "document_type": "Unknown", "confidence": 0.0, "reasoning": f"Error: {str(e)}"}


# ── TOOL: Analyse file content ─────────────────────────────────────────────────
def analyse_content_match(content: str, filename: str, search_context: str) -> dict:
    """Use LLM to assess whether the file's content confirms it matches the search context."""
    prompt = (
        f"You are a document classification assistant.\n\n"
        f"Search context — the user is looking for:\n{search_context}\n\n"
        f"Filename: {filename}\n"
        f"File content extract:\n---\n{content}\n---\n\n"
        f"Analyse the file CONTENTS. Determine:\n"
        f"1. Does the content confirm this is one of the documents being searched for?\n"
        f"2. What document type is it (be specific)?\n"
        f"3. Confidence score (0.0 = definitely not, 1.0 = definitely yes)\n\n"
        f'Respond only in JSON: {{"match": true/false, "document_type": "...", "confidence": 0.0, "reasoning": "..."}}'
    )
    try:
        return _llm_call(prompt)
    except Exception as e:
        return {"match": False, "document_type": "Unknown", "confidence": 0.0, "reasoning": f"Error: {str(e)}"}


# ── TOOL: Extract document metadata ───────────────────────────────────────────
def extract_document_metadata(content: str, filename: str) -> dict:
    """Use LLM to extract structured metadata from a document's content."""
    content_is_minimal = (
        not content or
        content.startswith("[") or
        len(content.strip()) < 100
    )
    content_note = (
        "NOTE: The file content could not be extracted (image-based or unreadable). "
        "Infer as much metadata as possible from the filename alone.\n\n"
        if content_is_minimal else ""
    )
    prompt = (
        f"You are a document metadata extraction assistant.\n\n"
        f"Filename: {filename}\n"
        f"{content_note}"
        f"File content extract:\n---\n{content}\n---\n\n"
        f"Extract the following metadata fields. When content is unavailable, infer from the filename "
        f"(e.g. project numbers, revision codes, document type keywords, dates encoded in filename). "
        f"If a field truly cannot be determined, use null.\n\n"
        f"Fields to extract:\n"
        f"- document_title: The full title of the document\n"
        f"- document_type: Specific document type (e.g. Design Brief, Structural Report, Drawing Register, TAL, Reticulation Drawing)\n"
        f"- author: Author name or organisation that produced the document\n"
        f"- date: Document date or revision date (as a string, e.g. '2024-11-12' or 'November 2024')\n"
        f"- project_name: Name of the project this document relates to\n"
        f"- project_number: Project number or job number if present\n"
        f"- revision: Document revision or version number if present\n\n"
        f'Respond only in JSON: {{"document_title": "...", "document_type": "...", "author": "...", "date": "...", "project_name": "...", "project_number": "...", "revision": "..."}}'
    )
    try:
        return _llm_call(prompt)
    except Exception as e:
        return {
            "document_title": None, "document_type": None, "author": None,
            "date": None, "project_name": None, "project_number": None, "revision": None
        }


# ── REQUEST / RESPONSE MODELS ──────────────────────────────────────────────────
class ReviewRequest(BaseModel):
    folder_path: str
    search_context: str
    output_dir: Optional[str] = None

class ProcessingPlanRequest(BaseModel):
    step1_output_file: str
    process_contexts: Optional[Dict[str, Any]] = None  # keyed by process_id, values are process document dicts
    output_dir: Optional[str] = None


# ── ENDPOINTS ──────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok", "agent": "Document Reviewer", "model": MODEL}


@app.post("/review")
def review_documents(request: ReviewRequest):
    # ── Tool: list_folder_files — ALL files (no extension filter) ───────────
    print(f"[Reviewer] Listing files in: {request.folder_path}")
    listing = _list_files_tool.run(folder_path=request.folder_path)

    if "error" in listing:
        raise HTTPException(status_code=400, detail=listing["error"])

    all_entries  = listing.get("files", [])
    file_entries = [e for e in all_entries if e["extension"] in SUPPORTED_EXTENSIONS]

    if not all_entries:
        return {
            "search_context": request.search_context,
            "folder_path": request.folder_path,
            "total_files_scanned": 0,
            "matched_files": 0,
            "results": []
        }

    scan_results = []

    # Add unsupported-extension files directly to scan_results (no processing)
    for entry in all_entries:
        if entry["extension"] not in SUPPORTED_EXTENSIONS:
            scan_results.append({
                "filename": entry["filename"],
                "file_size_bytes": entry.get("size_bytes"),
                "last_modified": entry.get("last_modified"),
                "content_type": "unsupported",
                "text_quality": "none",
                "recommended_reader": "unsupported",
                "quality_note": f"File type {entry['extension']} is not supported for text extraction.",
                "filename_match": False,
                "filename_confidence": 0.0,
                "filename_reasoning": "Unsupported file type",
                "content_match": False,
                "content_confidence": 0.0,
                "content_reasoning": "Unsupported file type",
                "matched_document_type": "Unknown",
                "overall_confidence": 0.0,
                "metadata": None
            })

    for entry in file_entries:
        file = Path(request.folder_path) / entry["filename"]
        print(f"[Reviewer] Checking filename: {entry['filename']}")

        # ── Tool 1: Filename check ──────────────────────────────────────────
        filename_result = check_filename_match(entry["filename"], request.search_context)

        # ── Tool 2: Read content (reused for both match and metadata) ───────
        print(f"[Reviewer] Reading content: {entry['filename']}")
        content = read_file_content(str(file))
        quality = assess_content_quality(content, Path(entry["filename"]).suffix)

        # ── Tool 3: Content match analysis ──────────────────────────────────
        content_result = analyse_content_match(content, entry["filename"], request.search_context)

        # ── Weighted overall confidence (filename 30%, content 70%) ─────────
        overall_confidence = round(
            filename_result.get("confidence", 0.0) * 0.3 +
            content_result.get("confidence", 0.0) * 0.7,
            2
        )

        # Prefer the content-derived document type, fall back to filename-derived
        document_type = (
            content_result.get("document_type") or
            filename_result.get("document_type") or
            "Unknown"
        )

        is_matched = content_result.get("match", False) or filename_result.get("match", False)

        # ── Tool 4: Extract metadata for matched documents ───────────────────
        metadata = None
        if is_matched:
            print(f"[Reviewer] Extracting metadata: {entry['filename']}")
            metadata = extract_document_metadata(content, entry["filename"])

        scan_results.append({
            "filename": entry["filename"],
            "file_size_bytes": entry.get("size_bytes"),
            "last_modified": entry.get("last_modified"),
            "content_type": quality["content_type"],
            "text_quality": quality["text_quality"],
            "recommended_reader": quality["recommended_reader"],
            "quality_note": quality["quality_note"],
            "filename_match": filename_result.get("match", False),
            "filename_confidence": round(filename_result.get("confidence", 0.0), 2),
            "filename_reasoning": filename_result.get("reasoning", ""),
            "content_match": content_result.get("match", False),
            "content_confidence": round(content_result.get("confidence", 0.0), 2),
            "content_reasoning": content_result.get("reasoning", ""),
            "matched_document_type": document_type,
            "overall_confidence": overall_confidence,
            "metadata": metadata
        })

    # Sort highest confidence first
    scan_results.sort(key=lambda x: x["overall_confidence"], reverse=True)

    matched = [r for r in scan_results if r["content_match"] or r["filename_match"]]

    # ── Build structured documents list for the next step ───────────────────────
    documents = []
    for r in matched:
        meta = r.get("metadata") or {}
        documents.append({
            "filename": r["filename"],
            "document_title": meta.get("document_title") or r["matched_document_type"],
            "document_type": meta.get("document_type") or r["matched_document_type"],
            "author": meta.get("author"),
            "date": meta.get("date"),
            "project_name": meta.get("project_name"),
            "project_number": meta.get("project_number"),
            "revision": meta.get("revision"),
            "confidence": r["overall_confidence"],
            "content_type": r["content_type"],
            "text_quality": r["text_quality"],
            "recommended_reader": r["recommended_reader"],
            "quality_note": r["quality_note"]
        })

    # ── Save output file ─────────────────────────────────────────────────────────
    out_dir = Path(request.output_dir) if request.output_dir else Path("/app/OUTPUT")
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        out_dir = Path("/app/OUTPUT")
        out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"DocumentReview_{timestamp}.json"
    output_path = out_dir / output_filename

    payload = {
        "search_context": request.search_context,
        "folder_path": request.folder_path,
        "total_files_scanned": len(all_entries),
        "matched_files": len(matched),
        "documents": documents,
        "scan_results": scan_results
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"[Reviewer] Output saved: {output_path}")

    return {
        **payload,
        "output_file": str(output_path),
        "files": {
            "files_read": [{"filename": e["filename"], "path": str(Path(request.folder_path) / e["filename"])} for e in all_entries],
            "files_output": [{"filename": output_filename, "path": str(output_path), "role": "document_review", "description": "Document review results with metadata and quality assessment"}]
        }
    }


# ── TOOL: Determine processing plan for a document ────────────────────────────
def _detect_page_size(filepath: str) -> dict:
    """Read PDF page dimensions and classify into standard paper sizes.
    Returns dict with page_size, page_width_mm, page_height_mm, requires_chunking, chunk_strategy, estimated_chunks.
    """
    # Paper size thresholds (mm, portrait orientation — compare short side × long side)
    # Allow ±10mm tolerance
    SIZE_MAP = [
        ("A4",  210, 297),
        ("A3",  297, 420),
        ("A2",  420, 594),
        ("A1",  594, 841),
        ("A0",  841, 1189),
    ]
    CHUNK_MAP = {"A2": 4, "A1": 6, "A0": 9, "large-format": 12}

    try:
        import pdfplumber
        with pdfplumber.open(filepath) as pdf:
            if not pdf.pages:
                return _default_page_size_info()
            page = pdf.pages[0]
            # pdfplumber gives dimensions in points (1pt = 0.352778mm)
            w_mm = round(page.width  * 0.352778)
            h_mm = round(page.height * 0.352778)
    except Exception:
        return _default_page_size_info()

    short = min(w_mm, h_mm)
    long_ = max(w_mm, h_mm)

    page_size = "large-format"
    for name, s, l in SIZE_MAP:
        if abs(short - s) <= 15 and abs(long_ - l) <= 15:
            page_size = name
            break

    requires_chunking = page_size in CHUNK_MAP
    chunk_strategy    = "quadrant-split" if requires_chunking else "none"
    estimated_chunks  = CHUNK_MAP.get(page_size, 1)

    return {
        "page_size":         page_size,
        "page_width_mm":     w_mm,
        "page_height_mm":    h_mm,
        "requires_chunking": requires_chunking,
        "chunk_strategy":    chunk_strategy,
        "estimated_chunks":  estimated_chunks,
    }


def _default_page_size_info() -> dict:
    return {
        "page_size":         "A4",
        "page_width_mm":     210,
        "page_height_mm":    297,
        "requires_chunking": False,
        "chunk_strategy":    "none",
        "estimated_chunks":  1,
    }


SPREADSHEET_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.xlsb'}

def _inspect_spreadsheet(filepath: str) -> dict:
    """Open a spreadsheet and return worksheet names, row counts, and non-empty column counts."""
    ext = Path(filepath).suffix.lower()
    try:
        if ext in {'.xlsx', '.xlsm', '.xlsb'}:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                row_count = ws.max_row or 0
                col_count = ws.max_column or 0
                sheets.append({"sheet_name": name, "row_count": row_count, "column_count": col_count})
            wb.close()
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(filepath)
            sheets = [
                {"sheet_name": sh.name, "row_count": sh.nrows, "column_count": sh.ncols}
                for sh in wb.sheets()
            ]
        else:
            return {"worksheet_count": 0, "worksheets": [], "spreadsheet_error": "Unsupported extension"}
        non_empty = [s for s in sheets if s["row_count"] > 1]
        return {
            "worksheet_count": len(sheets),
            "worksheets_to_analyse": len(non_empty),
            "worksheets": sheets,
        }
    except Exception as e:
        return {"worksheet_count": 0, "worksheets": [], "spreadsheet_error": str(e)}


def _determine_processing_plan(doc: dict, scan: dict, registry: dict, categorization_rules: str = "", page_size_info: dict = None) -> dict:
    """Use LLM to determine the best processing tool and approach for a document."""
    tools_summary = json.dumps(
        [
            {
                "id": t["id"],
                "name": t["name"],
                "description": t["description"],
                "use_case": t.get("use_case", ""),
            }
            for t in registry.get("tools", [])
        ],
        indent=2,
    )

    cat_section = ""
    cat_output_field = ""
    if categorization_rules:
        cat_section = (
            f"\nDocument categorisation rules (read these before assigning document_category):\n"
            f"{categorization_rules}\n"
        )
        cat_output_field = (
            f"6. document_category: Classify the document as EXACTLY 'Design Brief' or 'Site Plan' "
            f"using the categorisation rules above. Apply filename-based rules first (they take priority), "
            f"then use document_type / content as a fallback. Do NOT use any other value.\n"
        )

    psi = page_size_info or {}
    page_size       = psi.get("page_size", "unknown")
    page_w          = psi.get("page_width_mm", "unknown")
    page_h          = psi.get("page_height_mm", "unknown")
    req_chunk       = psi.get("requires_chunking", False)
    chunk_strat     = psi.get("chunk_strategy", "none")
    est_chunks      = psi.get("estimated_chunks", 1)

    response_fields = (
        '"processing_tool_id": "...", "conversion_required": true, '
        '"conversion_type": "..." or null, "extraction_method": "...", "processing_notes": "..."'
    )
    if categorization_rules:
        response_fields += ', "document_category": "Design Brief" or "Site Plan"'
    response_fields += (
        ', "page_size": "A4|A3|A2|A1|A0|large-format"'
        ', "requires_chunking": true/false'
        ', "chunk_strategy": "none|quadrant-split|page-split"'
        ', "estimated_chunks": integer'
    )

    prompt = (
        f"You are a document processing planner.\n\n"
        f"Given the following document information, determine the best processing approach.\n\n"
        f"Document:\n"
        f"- Filename: {doc.get('filename')}\n"
        f"- Document Type: {doc.get('document_type')}\n"
        f"- Content Type: {doc.get('content_type')} (text = extractable text, image = scanned/image-based)\n"
        f"- Text Quality: {doc.get('text_quality')} (high = >500 chars extracted, low = limited text, none = no text)\n"
        f"- Recommended Reader: {doc.get('recommended_reader')}\n"
        f"- Confidence Score: {doc.get('confidence')}\n"
        f"\nPage dimensions (measured from PDF):\n"
        f"- Page Size: {page_size} ({page_w}mm × {page_h}mm)\n"
        f"- Requires Chunking: {req_chunk} (pre-assessed: pages larger than A3 must be split for LLM processing)\n"
        f"- Suggested Chunk Strategy: {chunk_strat}\n"
        f"- Estimated Chunks per Page: {est_chunks}\n"
        + cat_section
        + f"\nAvailable tools in registry:\n{tools_summary}\n\n"
        f"Determine:\n"
        f"1. processing_tool_id: Which tool ID from the registry is best for reading/extracting this document\n"
        f"2. conversion_required: true/false — Does this document need format conversion before processing?\n"
        f"   (e.g. a scanned image-based PDF needs conversion to PNG/TIFF so an OCR tool can process it)\n"
        f"3. conversion_type: If conversion is required, what type? e.g. 'pdf-to-png', 'pdf-to-tiff', null if not required\n"
        f"4. extraction_method: Brief description of the extraction approach (e.g. 'Direct text extraction via pdfplumber', 'OCR after PDF-to-PNG conversion')\n"
        f"5. processing_notes: Any caveats or important notes for processing this document\n"
        + cat_output_field
        + f"6. page_size: Confirm the page size (use the measured value above unless there is clear reason to override)\n"
        f"7. requires_chunking: true if page_size is A2, A1, A0, or large-format; otherwise false\n"
        f"8. chunk_strategy: 'quadrant-split' for large drawings, 'page-split' for multi-page text docs with oversized pages, 'none' if not chunking\n"
        f"9. estimated_chunks: Expected number of chunks per page (A2→4, A1→6, A0→9, large-format→12, otherwise 1)\n"
        + f"\nRespond only in JSON: {{{response_fields}}}"
    )

    try:
        result = _llm_call(prompt)
        # Ensure page size fields are always present (use measured values as ground truth)
        result.setdefault("page_size",         psi.get("page_size", "A4"))
        result.setdefault("requires_chunking", psi.get("requires_chunking", False))
        result.setdefault("chunk_strategy",    psi.get("chunk_strategy", "none"))
        result.setdefault("estimated_chunks",  psi.get("estimated_chunks", 1))
        # Override with measured values if chunking is required but LLM missed it
        if psi.get("requires_chunking") and not result.get("requires_chunking"):
            result["requires_chunking"] = True
            result["chunk_strategy"]    = psi.get("chunk_strategy", "quadrant-split")
            result["estimated_chunks"]  = psi.get("estimated_chunks", 4)
        # RULE 0 enforced in code — spreadsheet extensions are always TAL, regardless of LLM output
        filename = doc.get("filename", "")
        if Path(filename).suffix.lower() in SPREADSHEET_EXTENSIONS:
            result["document_category"] = "TAL"
        # Assign route_to_step so downstream steps know which extraction step handles this file:
        #   TAL  → Step 6 (asset spreadsheet extraction, no chunking needed)
        #   Site Plan / Design Brief → Step 3 (chunk if required) then Step 4 or 5
        cat = result.get("document_category", "")
        if cat == "TAL":
            result["route_to_step"] = 6
            result["requires_chunking"] = False
            result["chunk_strategy"]    = "none"
            result["estimated_chunks"]  = 1
        elif cat == "Site Plan":
            result["route_to_step"] = 5
        elif cat == "Design Brief":
            result["route_to_step"] = 4
        else:
            result["route_to_step"] = None
        return result
    except Exception as e:
        filename = doc.get("filename", "")
        fallback_category = "TAL" if Path(filename).suffix.lower() in SPREADSHEET_EXTENSIONS else None
        result = {
            "processing_tool_id": "tool-read-file-content",
            "conversion_required": False,
            "conversion_type": None,
            "extraction_method": "Direct text extraction (fallback)",
            "processing_notes": f"Error during planning: {str(e)}",
            "page_size":         psi.get("page_size", "A4"),
            "requires_chunking": psi.get("requires_chunking", False),
            "chunk_strategy":    psi.get("chunk_strategy", "none"),
            "estimated_chunks":  psi.get("estimated_chunks", 1),
        }
        if fallback_category:
            result["document_category"] = fallback_category
        cat = result.get("document_category", "")
        if cat == "TAL":
            result["route_to_step"] = 6
            result["requires_chunking"] = False
            result["chunk_strategy"]    = "none"
            result["estimated_chunks"]  = 1
        elif cat == "Site Plan":
            result["route_to_step"] = 5
        elif cat == "Design Brief":
            result["route_to_step"] = 4
        else:
            result["route_to_step"] = None
        return result


# ── ENDPOINT: Plan document processing ────────────────────────────────────────
@app.post("/plan_processing")
def plan_document_processing(request: ProcessingPlanRequest):
    step1_file = Path(request.step1_output_file)
    if not step1_file.exists():
        raise HTTPException(status_code=400, detail=f"Step 1 output file not found: {request.step1_output_file}")

    print(f"[Reviewer] Loading Step 1 output: {step1_file}")
    with open(step1_file, "r", encoding="utf-8") as f:
        step1_data = json.load(f)

    # Load registry for available tools
    registry_path = Path(__file__).parent / "registry.json"
    registry: dict = {}
    if registry_path.exists():
        with open(registry_path, "r", encoding="utf-8") as f:
            registry = json.load(f)

    # Build categorization rules string from process contexts (loaded prior to execution)
    categorization_rules = ""
    if request.process_contexts:
        rules_parts = []
        for proc_id, proc_data in request.process_contexts.items():
            dc = proc_data.get("document_categorisation", {})
            if not dc:
                continue
            category    = dc.get("category", "")
            pos         = dc.get("positive_indicators", {})
            neg         = dc.get("negative_indicators", {})
            decision    = dc.get("decision_rule", "")
            rules_parts.append(
                f"Category '{category}':\n"
                f"  Positive filename patterns: {pos.get('filename_patterns', [])}\n"
                f"  Document type values: {pos.get('document_type_values', [])}\n"
                f"  Negative filename patterns (exclude): {neg.get('filename_patterns', [])}\n"
                f"  Decision rule: {decision}"
            )
        categorization_rules = "\n\n".join(rules_parts)
        if categorization_rules:
            print(f"[Reviewer] Loaded categorisation rules for {len(rules_parts)} process(es).")

    documents = step1_data.get("documents", [])
    scan_lookup = {r["filename"]: r for r in step1_data.get("scan_results", [])}

    processing_plan = []
    for doc in documents:
        filename = doc.get("filename", "")
        scan = scan_lookup.get(filename, {})
        print(f"[Reviewer] Planning processing for: {filename}")

        # Detect actual page dimensions / worksheet info before calling the LLM planner
        docs_folder = step1_data.get("folder_path", "/documents")
        doc_filepath = str(Path(docs_folder) / filename)
        file_ext = Path(filename).suffix.lower()

        if file_ext in SPREADSHEET_EXTENSIONS:
            page_size_info = _default_page_size_info()
            spreadsheet_info = _inspect_spreadsheet(doc_filepath)
            ws_count = spreadsheet_info.get("worksheets_to_analyse", 0)
            print(f"[Reviewer] Spreadsheet {filename}: {spreadsheet_info.get('worksheet_count')} sheet(s), {ws_count} to analyse")
        else:
            page_size_info = _detect_page_size(doc_filepath) if file_ext == ".pdf" else _default_page_size_info()
            spreadsheet_info = {}
            if page_size_info.get("page_size") != "A4":
                print(f"[Reviewer] Page size for {filename}: {page_size_info['page_size']} ({page_size_info['page_width_mm']}mm × {page_size_info['page_height_mm']}mm), chunking={page_size_info['requires_chunking']}")

        plan_entry = _determine_processing_plan(doc, scan, registry, categorization_rules, page_size_info)

        processing_plan.append({
            "filename": filename,
            "document_type": doc.get("document_type"),
            "content_type": doc.get("content_type"),
            "text_quality": doc.get("text_quality"),
            "recommended_reader": doc.get("recommended_reader"),
            "quality_note": doc.get("quality_note"),
            "confidence": doc.get("confidence"),
            **plan_entry,
            **spreadsheet_info,   # worksheet_count, worksheets_to_analyse, worksheets list
        })

    # Save output
    out_dir = Path(request.output_dir) if request.output_dir else Path("/app/OUTPUT")
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        out_dir = Path("/app/OUTPUT")
        out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"ProcessingPlan_{timestamp}.json"
    output_path = out_dir / output_filename

    payload = {
        "step1_output_file": str(step1_file),
        "total_documents": len(documents),
        "processing_plan": processing_plan,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    print(f"[Reviewer] Processing plan saved: {output_path}")

    return {
        **payload,
        "output_file": str(output_path),
        "files": {
            "files_read": [{"filename": step1_file.name, "path": str(step1_file)}],
            "files_output": [
                {
                    "filename": output_filename,
                    "path": str(output_path),
                    "role": "processing_plan",
                    "description": "Document processing plan with tool selection and conversion requirements per document",
                }
            ],
        },
    }
