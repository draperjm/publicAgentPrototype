"""
Document Extraction Agent
=========================
Receives a list of files (with processing metadata) and a process step definition.
Extracts content from each file using the appropriate tool, tags each section for
relevance to the process step, and returns a structured extraction JSON per file.

Endpoint: POST /extract
Port:     8090
"""

import concurrent.futures
import json
import logging
import os
import re
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import base64
import io

import anthropic
import google.generativeai as genai
import openai
import PIL.Image
import pdfplumber
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

logger = logging.getLogger("document_extractor")
logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Document Extraction Agent")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

LLM_MODEL              = os.getenv("LLM_MODEL",              "gpt-4o")
TAGGING_MODEL          = os.getenv("TAGGING_MODEL",          "gpt-4o-mini")
VISION_MODEL           = os.getenv("VISION_MODEL",           "gemini-2.0-flash")
ASSET_EXTRACTION_MODEL = os.getenv("ASSET_EXTRACTION_MODEL", "claude-sonnet-4-6")

_openai_client    = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
_anthropic_client = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
genai.configure(api_key=os.getenv("GOOGLE_API_KEY"))

# Per-request model routing — set in extract_documents, read by _llm_call
_req_context = threading.local()

OUTPUT_DIR = Path("/app/OUTPUT")
OUTPUT_DIR.mkdir(exist_ok=True)

# Sections per batch for tagging — smaller batches are more reliable
TAGGING_BATCH_SIZE = 15

# Retries and backoff for LLM calls
LLM_MAX_RETRIES  = 3
LLM_RETRY_DELAYS = [2, 5, 10]  # seconds between attempts

# Minimum alpha-character ratio to consider a section text meaningful
MIN_ALPHA_RATIO = 0.30

# Maximum sections before falling back to page-level grouping
MAX_SECTIONS_BEFORE_FALLBACK = 80


# ── Section heading patterns ──────────────────────────────────────────────────
# Numbered headings: "1.2 Section Title"
_NUMBERED = re.compile(
    r"^\s*(\d+(?:\.\d+){0,3})\s+([A-Z][A-Za-z0-9 ,\-/&:()\[\]']{2,70})\s*$"
)
# ALL-CAPS headings: must be at least 2 words, ≥8 chars, no repeated single word
# Excludes lines that are pure coordinates, addresses, or label spam
_ALLCAPS = re.compile(r"^\s*([A-Z][A-Z0-9]{1,}(?:\s+[A-Z][A-Z0-9]{1,}){1,})\s*$")

# High-value headings that always force a section boundary regardless of capitalisation.
# These are critical financial/legal sections commonly missed by the ALLCAPS detector.
_FORCED_HEADINGS = re.compile(
    r"^\s*(determination\s+of\s+(funding|supply)|ancillary\s+network\s+services(\s+fees?)?|"
    r"land\s+interests?|method\s+of\s+supply|funding\s+requirements?|"
    r"capital\s+contribution|contestable\s+works|non.contestable\s+works)\s*$",
    re.IGNORECASE,
)


def _is_valid_heading(text: str) -> bool:
    """Extra validation for ALL-CAPS candidate headings."""
    words = text.strip().split()
    if len(words) < 2 or len(text.strip()) < 8:
        return False
    # Reject if more than 60% of words are identical (e.g. "GAS GAS GAS GAS")
    if len(set(words)) / len(words) < 0.5:
        return False
    # Reject if it looks like a street address (contains NSW, VIC, QLD, etc.)
    address_tokens = {"NSW", "VIC", "QLD", "SA", "WA", "TAS", "NT", "ACT"}
    if address_tokens & set(words):
        return False
    # Reject if it looks like a date (contains month names or years)
    months = {"JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE","JULY",
              "AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"}
    if months & set(words):
        return False
    return True


def _alpha_ratio(text: str) -> float:
    """Return fraction of characters that are alphabetic."""
    if not text:
        return 0.0
    alpha = sum(c.isalpha() for c in text)
    return alpha / len(text)


_JSON_FENCE = re.compile(r"```(?:json)?\s*([\s\S]*?)```", re.IGNORECASE)

def _parse_json_robust(text: str) -> Any:
    """Parse JSON from LLM response, stripping markdown code fences if present."""
    text = text.strip()
    m = _JSON_FENCE.search(text)
    if m:
        text = m.group(1).strip()
    return json.loads(text)


def _llm_call_claude(prompt: str, model: str) -> Any:
    """Call a Claude model via Anthropic API. Returns parsed JSON."""
    last_exc = None
    for attempt in range(LLM_MAX_RETRIES):
        try:
            msg = _anthropic_client.messages.create(
                model=model,
                max_tokens=16000,
                temperature=0,
                messages=[{"role": "user", "content": prompt}],
            )
            return _parse_json_robust(msg.content[0].text)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Claude call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call(prompt: str) -> Any:
    """Call the configured LLM with retry logic. Routes to Claude or OpenAI based on request context."""
    active_model = getattr(_req_context, "model", LLM_MODEL)
    if active_model.startswith("claude-"):
        return _llm_call_claude(prompt, active_model)
    last_exc = None
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = _openai_client.chat.completions.create(
                model=active_model,
                messages=[{"role": "user", "content": prompt}],
                response_format={"type": "json_object"},
                temperature=0,
            )
            return _parse_json_robust(resp.choices[0].message.content)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"LLM call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call_fast(prompt: str) -> Any:
    """Call GPT-4o-mini (fast tagging) with retry logic. Returns parsed JSON."""
    last_exc = None
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = _openai_client.chat.completions.create(
                model=TAGGING_MODEL,
                messages=[{"role": "user", "content": prompt}],
                response_format={"type": "json_object"},
                temperature=0,
            )
            return _parse_json_robust(resp.choices[0].message.content)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Fast LLM call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


def _llm_call_vision(prompt: str, image_bytes: bytes) -> Any:
    """Call Gemini with an inline PNG image + prompt. Returns parsed JSON."""
    img = PIL.Image.open(io.BytesIO(image_bytes))
    last_exc = None
    model = genai.GenerativeModel(VISION_MODEL)
    for attempt in range(LLM_MAX_RETRIES):
        try:
            resp = model.generate_content([prompt, img])
            return _parse_json_robust(resp.text)
        except Exception as e:
            last_exc = e
            delay = LLM_RETRY_DELAYS[min(attempt, len(LLM_RETRY_DELAYS) - 1)]
            logger.warning(f"Vision LLM call failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}. Retrying in {delay}s…")
            time.sleep(delay)
    raise last_exc


# ── Request / response models ─────────────────────────────────────────────────

class FileEntry(BaseModel):
    filename: str
    filepath: Optional[str] = None
    processing_tool_id: Optional[str] = "tool-extract-pdf-content"
    document_type: Optional[str] = None
    document_category: Optional[str] = None
    content_type: Optional[str] = "text"
    text_quality: Optional[str] = "high"
    page_size: Optional[str] = "A4"
    requires_chunking: Optional[bool] = False
    chunk_strategy: Optional[str] = "none"
    estimated_chunks: Optional[int] = 1
    chunk_manifest: Optional[Dict[str, Any]] = None  # populated by chunker step


class ProcessStepDef(BaseModel):
    step_id: Optional[str] = None
    step_name: str
    summary: Optional[str] = None
    details: Optional[str] = None
    expected_output: Optional[Dict[str, Any]] = None
    sub_steps: Optional[List[Dict[str, Any]]] = None
    instructional_sub_steps: Optional[List[Dict[str, Any]]] = None


class ExtractionRequest(BaseModel):
    files: List[FileEntry]
    process_step: ProcessStepDef
    documents_folder: Optional[str] = "/documents"
    output_dir: Optional[str] = None
    process_id: Optional[str] = None   # e.g. "proc-extract-asset-spreadsheet"


# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "service": "document-extractor", "model": LLM_MODEL}


class RawTextRequest(BaseModel):
    filepaths: List[str]
    documents_folder: Optional[str] = "/documents"


@app.post("/raw_text")
def extract_raw_text(request: RawTextRequest):
    """
    Return full page-by-page raw text from one or more PDFs.
    Used by the orchestrator to feed source content to the step validator.
    """
    results = {}
    for fp in request.filepaths:
        path = fp if Path(fp).is_absolute() else str(Path(request.documents_folder) / fp)
        filename = Path(path).name
        try:
            with pdfplumber.open(path) as pdf:
                pages = []
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    raw_tables = page.extract_tables() or []
                    page_tables = []
                    for tbl in raw_tables:
                        if tbl and len(tbl) > 1:
                            page_tables.append({
                                "headers": [str(h or "") for h in tbl[0]],
                                "rows":    [[str(c or "") for c in row] for row in tbl[1:]],
                            })
                    pages.append({"page": page.page_number, "text": text, "tables": page_tables})
            full_text = "\n\n".join(
                f"--- Page {p['page']} ---\n{p['text']}" for p in pages if p["text"].strip()
            )
            results[filename] = {
                "filepath": path, "total_pages": len(pages),
                "pages": pages, "full_text": full_text, "error": None,
            }
        except Exception as e:
            results[filename] = {
                "filepath": path, "total_pages": 0,
                "pages": [], "full_text": "", "error": str(e),
            }
    return {"documents": results}


@app.post("/extract")
def extract_documents(request: ExtractionRequest):
    """
    Extract and tag content from each file according to the process step.
    Returns a consolidated extraction report and saves per-file JSON files.
    """
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    extractions = []

    # Route _llm_call to the appropriate model for this process
    if request.process_id == "proc-extract-asset-spreadsheet":
        _req_context.model = ASSET_EXTRACTION_MODEL
        logger.info(f"Asset extraction — using model: {ASSET_EXTRACTION_MODEL}")
    else:
        _req_context.model = LLM_MODEL

    # Use job-scoped output directory when provided, else fall back to module-level OUTPUT_DIR
    eff_output_dir = Path(request.output_dir) if request.output_dir else OUTPUT_DIR
    try:
        eff_output_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        eff_output_dir = OUTPUT_DIR
        eff_output_dir.mkdir(parents=True, exist_ok=True)

    for fe in request.files:
        filepath = fe.filepath or str(Path(request.documents_folder) / fe.filename)
        if request.process_id == "proc-extract-asset-spreadsheet":
            logger.info(f"Asset extraction file: filename={fe.filename!r}, filepath={fe.filepath!r}, "
                        f"documents_folder={request.documents_folder!r}, resolved={filepath!r}")
            result = _extract_asset_spreadsheet(fe, filepath)
        else:
            result = _extract_file(fe, filepath, request.process_step)

        # Save per-file output
        safe_name  = re.sub(r"[^\w]", "_", fe.filename)
        out_path   = eff_output_dir / f"Extraction_{safe_name}_{timestamp}.json"
        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)
        result["output_file"] = str(out_path)

        extractions.append(result)

    # Consolidated report
    report_path = eff_output_dir / f"ExtractionReport_{timestamp}.json"
    report = {
        "process_step": request.process_step.dict(),
        "documents_folder": request.documents_folder,
        "total_files": len(extractions),
        "extractions": extractions,
        "timestamp": timestamp,
        "output_file": str(report_path),
    }
    with open(report_path, "w") as f:
        json.dump(report, f, indent=2)

    # For asset spreadsheet extraction, also produce a flat AssetExtract JSON
    # consolidating all asset_records across all files for use in downstream steps.
    if request.process_id == "proc-extract-asset-spreadsheet":
        all_asset_records = []
        source_documents  = []
        for ex in extractions:
            se = ex.get("step_extraction") or {}
            records = se.get("asset_records") or []
            all_asset_records.extend(records)
            if ex.get("filename"):
                source_documents.append(ex["filename"])
        asset_extract_path = eff_output_dir / f"AssetExtract_{timestamp}.json"
        asset_extract = {
            "process_id":       "proc-extract-asset-spreadsheet",
            "timestamp":        timestamp,
            "source_documents": source_documents,
            "total_assets":     len(all_asset_records),
            "asset_records":    all_asset_records,
            "output_file":      str(asset_extract_path),
        }
        with open(asset_extract_path, "w") as f:
            json.dump(asset_extract, f, indent=2)
        report["asset_extract"]      = asset_extract
        report["asset_extract_file"] = str(asset_extract_path)
        logger.info(f"AssetExtract saved: {asset_extract_path} ({len(all_asset_records)} records)")

    return report


# ── Core extraction logic ─────────────────────────────────────────────────────

def _extract_file(fe: FileEntry, filepath: str, step: ProcessStepDef) -> dict:
    """Full extraction pipeline for a single document."""
    # If chunk PNGs are available, use vision-based extraction instead of PDF text parsing
    if fe.chunk_manifest and fe.chunk_manifest.get("chunks"):
        return _extract_chunked_file(fe, fe.chunk_manifest, step)

    tool_id = (fe.processing_tool_id or "").lower()

    # Determine page grouping for large documents.
    # For quadrant-split (A2/A1/A0 drawings), use group_size=1 so every page
    # becomes its own chunk — prevents large drawing pages from being merged.
    # For page-split, same treatment. For none, use default (let detector decide).
    chunk_override: Optional[int] = None
    if fe.requires_chunking:
        if fe.chunk_strategy in ("quadrant-split", "page-split"):
            chunk_override = 1  # one page per chunk

    # 1. Extract raw sections from the document
    if filepath.lower().endswith(".pdf") or "pdf" in tool_id:
        raw_sections, doc_quality = _extract_pdf_sections(filepath, chunk_page_size=chunk_override)
    else:
        raw_sections = _extract_text_file(filepath)
        doc_quality = "text"

    # 2. Tag all sections in batched LLM calls (chunked)
    tagged_sections = _tag_all_sections_chunked(raw_sections, step)

    # 3. Isolate relevant sections (score >= 0.4)
    relevant = [s for s in tagged_sections if s.get("relevance_score", 0) >= 0.4]

    base = {
        "filename":              fe.filename,
        "filepath":              filepath,
        "document_type":         fe.document_type,
        "document_category":     fe.document_category,
        "input_text_quality":    fe.text_quality,   # quality as assessed by planner
        "document_quality":      doc_quality,       # quality as computed from PDF content
        "page_size":             fe.page_size or "A4",
        "requires_chunking":     fe.requires_chunking or False,
        "chunk_strategy":        fe.chunk_strategy or "none",
        "estimated_chunks":      fe.estimated_chunks or 1,
        "process_step_name":     step.step_name,
        "total_sections":        len(raw_sections),
        "relevant_sections":     len(relevant),
    }

    # 4a. If sub_steps defined, run a targeted extraction per sub-step.
    # Pass all tagged sections so each sub-step can apply its own threshold.
    if step.sub_steps:
        base["sub_step_extractions"] = _extract_per_sub_step(tagged_sections, step.sub_steps)
    else:
        # 4b. Single structured extraction across all relevant sections
        base["step_extraction"] = _structured_extraction(relevant, step)

    return base


def _extract_chunked_file(fe: FileEntry, chunk_manifest: dict, step: ProcessStepDef) -> dict:
    """
    Vision-based extraction for documents that have been pre-split into PNG chunks.

    Each chunk PNG is sent to Gemini with the process step instructions.
    Results from all chunks are consolidated into a single structured output
    that matches the format produced by _extract_file.
    """
    chunks     = chunk_manifest.get("chunks", [])
    strategy   = chunk_manifest.get("chunk_strategy", fe.chunk_strategy or "quadrant-split")
    page_size  = chunk_manifest.get("page_size", fe.page_size or "A4")
    total_pages = chunk_manifest.get("total_pages", 1)

    # Build extraction instruction from the process step definition
    inst_parts = [f"Process step: {step.step_name}"]
    if step.summary:
        inst_parts.append(f"Summary: {step.summary}")
    if step.details:
        inst_parts.append(f"Details: {step.details}")
    if step.sub_steps:
        inst_parts.append("Extract the following fields:")
        for ss in step.sub_steps:
            inst_parts.append(
                f"  - {ss.get('sub_step_name', ss.get('sub_step_id', ''))}: {ss.get('details', '')}"
            )
    instruction = "\n".join(inst_parts)

    # ── Phase 1: process each chunk with the vision model ─────────────────────
    chunk_results: List[dict] = []
    for chunk in chunks:
        chunk_id   = chunk.get("chunk_id", "?")
        region     = chunk.get("region", "")
        page_num   = chunk.get("page_number", 1)
        fpath      = chunk.get("filepath", "")

        if not fpath or not Path(fpath).exists():
            logger.warning(f"Chunk file missing: {fpath}")
            chunk_results.append({"chunk_id": chunk_id, "region": region,
                                   "page_number": page_num, "extracted": None,
                                   "error": "file not found"})
            continue

        prompt = (
            f"You are an expert engineering document analyst.\n"
            f"This image is region '{region}' (chunk {chunk_id}, page {page_num} of {total_pages}) "
            f"from a {page_size} engineering document split using {strategy}.\n\n"
            f"{instruction}\n\n"
            f"Return a JSON object with:\n"
            f"  'raw_text': all legible text visible in this image chunk,\n"
            f"  'data': structured fields relevant to the process step extracted from this chunk "
            f"(use null for fields not visible here).\n"
            f"Only include information actually present in this image."
        )

        try:
            img_bytes = Path(fpath).read_bytes()
            result    = _llm_call_vision(prompt, img_bytes)
            chunk_results.append({
                "chunk_id":    chunk_id,
                "sequence":    chunk.get("sequence", 0),
                "page_number": page_num,
                "region":      region,
                "extracted":   result,
            })
        except Exception as e:
            logger.warning(f"Vision extraction failed for {chunk_id}: {e}")
            chunk_results.append({
                "chunk_id":    chunk_id,
                "sequence":    chunk.get("sequence", 0),
                "page_number": page_num,
                "region":      region,
                "extracted":   None,
                "error":       str(e),
            })

    # ── Phase 2: consolidate all chunk observations ────────────────────────────
    # Collect raw text and per-chunk data observations
    raw_text_parts: List[str] = []
    all_data_fragments: List[dict] = []

    for cr in chunk_results:
        ext = cr.get("extracted") or {}
        raw = ext.get("raw_text", "") if isinstance(ext, dict) else ""
        if raw:
            raw_text_parts.append(f"[{cr['region']}, p{cr['page_number']}] {raw}")
        data = ext.get("data") if isinstance(ext, dict) else None
        if isinstance(data, dict):
            all_data_fragments.append(data)

    combined_raw = "\n\n".join(raw_text_parts)
    successful   = [cr for cr in chunk_results if cr.get("extracted")]

    # Merge data fragments: first non-null value per key wins
    merged_data: dict = {}
    for frag in all_data_fragments:
        for k, v in frag.items():
            if v is not None and k not in merged_data:
                merged_data[k] = v

    # ── Phase 3: final consolidation LLM pass (if sub_steps present) ──────────
    # Feed all raw observations back to the model to produce the final structured output
    if step.sub_steps and combined_raw:
        consolidation_prompt = (
            f"You are an expert document analyst. The following text was extracted from all "
            f"regions of a {page_size} engineering document ({strategy} chunking).\n\n"
            f"EXTRACTED TEXT:\n{combined_raw[:8000]}\n\n"
            f"PARTIAL STRUCTURED DATA:\n{json.dumps(merged_data, indent=2)[:4000]}\n\n"
            f"{instruction}\n\n"
            f"Using all the above information, produce a final consolidated JSON extraction. "
            f"For each sub-step field, provide the best answer found across all chunks. "
            f"Return a JSON object with a key per sub-step id containing the extracted value."
        )
        try:
            consolidated = _llm_call(consolidation_prompt)
        except Exception as e:
            logger.warning(f"Consolidation LLM call failed: {e}. Using merged chunk data.")
            consolidated = merged_data
    else:
        consolidated = merged_data

    # ── Assemble output in the same shape as _extract_file ────────────────────
    base = {
        "filename":              fe.filename,
        "filepath":              fe.filepath or "",
        "document_type":         fe.document_type,
        "document_category":     fe.document_category,
        "input_text_quality":    fe.text_quality,
        "document_quality":      "chunked-image",
        "page_size":             page_size,
        "requires_chunking":     True,
        "chunk_strategy":        strategy,
        "estimated_chunks":      len(chunks),
        "actual_chunks_processed": len(successful),
        "process_step_name":     step.step_name,
        "total_sections":        len(chunks),
        "relevant_sections":     len(successful),
        "raw_text_from_chunks":  combined_raw,
        "chunk_details":         chunk_results,
    }

    if step.sub_steps:
        # Map each sub-step id to its value from consolidated output
        sub_step_extractions = {}
        for ss in step.sub_steps:
            ss_id = ss.get("sub_step_id") or f"sub_step_{ss.get('sub_step_number', '')}"
            sub_step_extractions[ss_id] = {
                "value":          consolidated.get(ss_id) if isinstance(consolidated, dict) else None,
                "sub_step_name":  ss.get("sub_step_name", ss_id),
            }
        base["sub_step_extractions"] = sub_step_extractions
    else:
        base["step_extraction"] = {"value": consolidated}

    return base


def _derive_tag_hint(sub_step_id: str) -> str:
    """Map a sub_step_id to a content_tag string used during section tagging."""
    tag_map = {
        "hv":          "hv",
        "lv":          "lv",
        "sl":          "sl",
        "earthing":    "earthing",
        "easement":    "easement",
        "substation":  "substation",
        "funding":     "funding",
        "transformer": "transformer",
        "voltage":     "voltage",
        "cubicle":     "cubicle",
        "switchgear":  "switchgear",
        "siteplan":    "siteplan",
    }
    sid_lower = sub_step_id.lower()
    for key, tag in tag_map.items():
        if key in sid_lower:
            return tag
    return ""


def _extract_per_sub_step(all_tagged: list, sub_steps: list) -> dict:
    """
    Run a separate targeted extraction for each sub-step.

    Per-sub-step behaviour:
    - Low-threshold categories (funding, easement) use score >= 0.25 so short
      checklist-style sections are not filtered out.
    - All other categories use score >= 0.4.
    - Sections are narrowed by content_tag match when possible.
    - When no tag match exists, fall back to the top 5 by relevance score only
      (not all sections) to prevent unrelated content flooding the prompt.
    - Checklist categories (easement, earthing) also force-include the first two
      document sections regardless of score, since these items often appear as
      short header lines near the top of the document.
    """
    # Categories with lower relevance threshold (short/sparse but critical content)
    LOW_THRESHOLD_TAGS = {"funding", "easement"}
    # Categories that additionally force-include the first two document sections
    CHECKLIST_TAGS = {"easement", "earthing"}

    # Build (ss_id, ss_relevant, ss_step) tuples for all sub-steps up front,
    # then run all structured extractions in parallel.
    tasks: List[tuple] = []
    for ss in sub_steps:
        ss_id      = ss.get("sub_step_id") or f"sub_step_{ss.get('sub_step_number', '')}"
        ss_name    = ss.get("sub_step_name", ss_id)
        ss_details = ss.get("details", "")
        ss_inst    = ss.get("instructional_sub_steps") or []

        ss_step = ProcessStepDef(
            step_id=ss_id,
            step_name=ss_name,
            details=ss_details,
            instructional_sub_steps=ss_inst if ss_inst else None,
        )

        tag_hint  = _derive_tag_hint(ss_id)
        threshold = 0.25 if tag_hint in LOW_THRESHOLD_TAGS else 0.4

        above_threshold = [s for s in all_tagged if s.get("relevance_score", 0) >= threshold]

        if tag_hint:
            tag_filtered = [s for s in above_threshold if tag_hint in s.get("content_tags", [])]
            if tag_filtered:
                ss_relevant = tag_filtered
            else:
                ss_relevant = sorted(
                    above_threshold,
                    key=lambda s: s.get("relevance_score", 0),
                    reverse=True,
                )[:5]
        else:
            ss_relevant = above_threshold

        if tag_hint in CHECKLIST_TAGS:
            first_two = all_tagged[:2]
            existing_names = {s.get("section_name") for s in ss_relevant}
            for s in first_two:
                if s.get("section_name") not in existing_names:
                    ss_relevant = [s] + ss_relevant

        tasks.append((ss_id, ss_relevant, ss_step))

    # Run all sub-step extractions concurrently — each is an independent LLM call.
    results = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=len(tasks)) as executor:
        future_to_id = {
            executor.submit(_structured_extraction, ss_rel, ss_stp): ss_id
            for ss_id, ss_rel, ss_stp in tasks
        }
        for future in concurrent.futures.as_completed(future_to_id):
            sid = future_to_id[future]
            try:
                results[sid] = future.result()
            except Exception as e:
                logger.error(f"Sub-step extraction failed for {sid}: {e}")
                results[sid] = {"error": str(e)}

    return results


# ── PDF extraction ────────────────────────────────────────────────────────────

def _extract_pdf_sections(filepath: str, chunk_page_size: Optional[int] = None):
    """
    Split a PDF into logical sections using heading detection.

    Returns (sections_list, quality_flag) where quality_flag is one of:
      'text'    — normal text PDF, section-level split used
      'sparse'  — low text density, fell back to page-level grouping
      'drawing' — very low alpha ratio (likely CAD/drawing), minimal text
      'chunked' — forced page-level split due to large page size

    Each section dict: {section_number, section_name, page_number, text, tables, images}
    chunk_page_size: when set, forces page-level grouping with this group size (e.g. 1 for large drawings).
    """
    try:
        with pdfplumber.open(filepath) as pdf:
            # ── Pass 1: collect all page text and tables ──────────────────────
            page_data = []
            for page in pdf.pages:
                text = page.extract_text() or ""
                tables = []
                for tbl in (page.extract_tables() or []):
                    if tbl and len(tbl) > 1:
                        tables.append({
                            "headers": [str(h or "") for h in tbl[0]],
                            "rows":    [[str(c or "") for c in row] for row in tbl[1:]],
                        })
                page_data.append({"page": page.page_number, "text": text, "tables": tables})

    except Exception as e:
        return ([{
            "section_number": None, "section_name": "Extraction Error",
            "page_number": 0, "text": str(e), "tables": [], "images": [],
        }], "error")

    # ── Force chunk-level split for large pages ───────────────────────────────
    if chunk_page_size is not None:
        sections = _page_level_sections(page_data, group_size=chunk_page_size)
        return sections, "chunked"

    # ── Assess overall document text quality ─────────────────────────────────
    all_text = "\n".join(p["text"] for p in page_data)
    overall_alpha = _alpha_ratio(all_text)

    # Drawing / CAD PDF: very low alpha ratio — return page-level sections
    if overall_alpha < 0.25:
        sections = _page_level_sections(page_data, group_size=2)
        return sections, "drawing"

    # ── Pass 2: heading-based section split ───────────────────────────────────
    sections: list = []
    current = _new_section(None, "Preamble", 1)

    for pd in page_data:
        pnum = pd["page"]
        for line in pd["text"].split("\n"):
            mn = _NUMBERED.match(line)
            mf = _FORCED_HEADINGS.match(line) if not mn else None
            mc = _ALLCAPS.match(line) if not mn and not mf else None

            if mn:
                _flush(sections, current)
                current = _new_section(mn.group(1), mn.group(2).strip(), pnum)
            elif mf:
                # High-value keyword heading — always split regardless of capitalisation
                _flush(sections, current)
                current = _new_section(None, line.strip(), pnum)
            elif mc and _is_valid_heading(mc.group(1)):
                _flush(sections, current)
                current = _new_section(None, mc.group(1).strip(), pnum)
            else:
                current["text"] += line + "\n"

        for tbl in pd["tables"]:
            current["tables"].append(tbl)

    _flush(sections, current)

    # ── Fallback: too many sections → page-level grouping ────────────────────
    if len(sections) > MAX_SECTIONS_BEFORE_FALLBACK:
        sections = _page_level_sections(page_data, group_size=3)
        return sections, "sparse"

    if not sections:
        sections = [{
            "section_number": None, "section_name": "Full Document",
            "page_number": 1, "text": all_text[:8000],
            "tables": [], "images": [],
        }]

    return sections, "text"


def _page_level_sections(page_data: list, group_size: int = 3) -> list:
    """
    Group consecutive pages into sections as a fallback for documents
    where heading detection produces too many or too few sections.
    """
    sections = []
    for i in range(0, len(page_data), group_size):
        chunk = page_data[i:i + group_size]
        combined_text = "\n\n".join(p["text"] for p in chunk if p["text"].strip())
        tables = [t for p in chunk for t in p["tables"]]
        if combined_text.strip() or tables:
            pstart = chunk[0]["page"]
            pend   = chunk[-1]["page"]
            name   = f"Pages {pstart}–{pend}" if pstart != pend else f"Page {pstart}"
            sections.append({
                "section_number": None, "section_name": name,
                "page_number": pstart, "text": combined_text,
                "tables": tables, "images": [],
            })
    return sections


def _new_section(num, name, page):
    return {"section_number": num, "section_name": name, "page_number": page,
            "text": "", "tables": [], "images": []}


def _flush(sections, current):
    text = current["text"].strip()
    # Only flush if the text has meaningful alphabetic content
    if (text and _alpha_ratio(text) >= MIN_ALPHA_RATIO) or current["tables"]:
        sections.append(dict(current))


# ── Spreadsheet extraction ────────────────────────────────────────────────────

_SPREADSHEET_EXTENSIONS = {'.xlsx', '.xls', '.xlsm', '.xlsb'}


def _read_spreadsheet_as_tables(filepath: str) -> list:
    """
    Read an Excel workbook and return a list of worksheet dicts:
        {"sheet_name": str, "headers": [...], "rows": [[...], ...]}
    Supports .xlsx/.xlsm/.xlsb via openpyxl and .xls via xlrd.
    """
    ext = Path(filepath).suffix.lower()
    sheets = []
    try:
        if ext in {'.xlsx', '.xlsm', '.xlsb'}:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = []
                for row in ws.iter_rows(values_only=True):
                    all_rows.append([str(c) if c is not None else "" for c in row])
                # Drop fully-empty rows
                all_rows = [r for r in all_rows if any(c.strip() for c in r)]
                if not all_rows:
                    continue
                headers = all_rows[0]
                data_rows = all_rows[1:]
                sheets.append({"sheet_name": sheet_name, "headers": headers, "rows": data_rows})
            wb.close()
        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for idx in range(wb.nsheets):
                ws  = wb.sheet_by_index(idx)
                name = wb.sheet_names()[idx]
                all_rows = []
                for i in range(ws.nrows):
                    all_rows.append([str(ws.cell_value(i, j)) for j in range(ws.ncols)])
                all_rows = [r for r in all_rows if any(c.strip() for c in r)]
                if not all_rows:
                    continue
                headers = all_rows[0]
                data_rows = all_rows[1:]
                sheets.append({"sheet_name": name, "headers": headers, "rows": data_rows})
    except Exception as e:
        logger.error(f"_read_spreadsheet_as_tables failed for {filepath}: {e}")
    return sheets


def _select_target_worksheets(sheets: list, filename: str) -> list:
    """
    Pass 1: Ask Claude to review all worksheet names + headers and return
    the list of sheet names that contain asset record data worth extracting.
    Returns a list of sheet names to process.
    """
    sheet_summaries = []
    for s in sheets:
        header_preview = "\t".join(s["headers"][:20])  # first 20 columns
        row_count      = len(s["rows"])
        # Show first 3 data rows as a sample
        sample_rows = "\n".join("\t".join(r[:20]) for r in s["rows"][:3])
        sheet_summaries.append(
            f"Sheet: {s['sheet_name']} ({row_count} data rows)\n"
            f"Headers: {header_preview}\n"
            f"Sample rows:\n{sample_rows}"
        )

    prompt = (
        f"You are reviewing a workbook named '{filename}' to identify which worksheets "
        f"contain asset register data suitable for extraction.\n\n"
        f"WORKSHEETS IN THIS FILE:\n\n"
        + "\n\n---\n\n".join(sheet_summaries)
        + "\n\n"
        "Asset register worksheets typically contain rows of individual assets with columns for "
        "identifiers (Asset ID, Tag, FLOC), asset type or description, location, status, "
        "technical specifications (voltage, kVA, capacity), manufacturer/model, condition, "
        "or installation/maintenance dates.\n\n"
        "Exclude worksheets that are: instructions/cover sheets, lookup tables, charts-only, "
        "summary/pivot tables, or empty/placeholder tabs.\n\n"
        "Return ONLY a JSON object: "
        "{\"target_sheets\": [\"SheetName1\", \"SheetName2\"], "
        "\"reasoning\": \"one sentence explaining the selection\"}"
    )

    try:
        result = _llm_call(prompt)
        targets = result.get("target_sheets") or []
        reasoning = result.get("reasoning", "")
        logger.info(f"Worksheet selection for '{filename}': {targets} — {reasoning}")
        # Validate: only return names that actually exist in the workbook
        valid_names = {s["sheet_name"] for s in sheets}
        return [t for t in targets if t in valid_names]
    except Exception as e:
        logger.warning(f"Worksheet selection failed: {e}. Processing all sheets.")
        return [s["sheet_name"] for s in sheets]


def _extract_asset_spreadsheet(fe: FileEntry, filepath: str) -> dict:
    """
    Dedicated extraction for proc-extract-asset-spreadsheet.
    Pass 1 — Claude selects which worksheets contain asset data.
    Pass 2 — Claude extracts structured asset records from each targeted sheet.
    Returns a result dict with step_extraction.asset_records.
    """
    logger.info(f"_extract_asset_spreadsheet: filepath={filepath!r}, exists={Path(filepath).exists()}")

    # If the explicit filepath doesn't exist, try to locate the file by name
    if not Path(filepath).exists():
        # Search common container mount points
        search_roots = [Path("/documents"), Path("/app/documents"), Path("/data")]
        found = None
        for root in search_roots:
            if root.exists():
                matches = list(root.rglob(fe.filename))
                if matches:
                    found = str(matches[0])
                    logger.info(f"Found spreadsheet at: {found}")
                    break
        if found:
            filepath = found
        else:
            logger.error(f"Spreadsheet not found: {filepath!r} (also searched /documents, /app/documents, /data)")
            return {
                "filename": fe.filename, "filepath": filepath,
                "document_type": fe.document_type, "document_category": fe.document_category,
                "document_quality": "spreadsheet",
                "step_extraction": {"asset_records": [], "total_assets": 0,
                                    "sheets_processed": [],
                                    "note": f"File not found: {filepath}"},
            }

    sheets = _read_spreadsheet_as_tables(filepath)
    if not sheets:
        return {
            "filename": fe.filename, "filepath": filepath,
            "document_type": fe.document_type, "document_category": fe.document_category,
            "document_quality": "spreadsheet",
            "step_extraction": {"asset_records": [], "total_assets": 0,
                                "sheets_processed": [], "note": "No readable worksheets found."},
        }

    # Pass 1: let Claude decide which worksheets to target
    target_names   = _select_target_worksheets(sheets, fe.filename)
    target_sheets  = [s for s in sheets if s["sheet_name"] in target_names]
    sheets_skipped = [s["sheet_name"] for s in sheets if s["sheet_name"] not in target_names]
    if sheets_skipped:
        logger.info(f"Skipping non-asset worksheets: {sheets_skipped}")

    # Fallback: if selection returned nothing, process all sheets
    if not target_sheets:
        target_sheets = sheets

    all_asset_records: list = []
    sheets_processed: list  = []

    # Pass 2: extract asset records from each targeted sheet
    for sheet in target_sheets:
        data_rows = sheet["rows"]
        if not data_rows:
            continue

        sheets_processed.append(sheet["sheet_name"])

        # Process rows in batches to stay within Claude's output token limit
        BATCH_SIZE  = 50
        header_line = "\t".join(sheet["headers"])
        sheet_records: list = []
        batches = [data_rows[i:i + BATCH_SIZE] for i in range(0, min(len(data_rows), 500), BATCH_SIZE)]

        for batch_idx, batch in enumerate(batches):
            row_lines  = ["\t".join(r) for r in batch]
            table_text = header_line + "\n" + "\n".join(row_lines)

            prompt = (
                f"You are extracting structured asset records from a spreadsheet worksheet.\n\n"
                f"Worksheet: {sheet['sheet_name']}\n"
                f"Source file: {fe.filename}\n"
                f"Batch: rows {batch_idx * BATCH_SIZE + 1}–{batch_idx * BATCH_SIZE + len(batch)} "
                f"of {len(data_rows)}\n\n"
                f"SPREADSHEET DATA (tab-separated, first row is header):\n"
                f"{table_text}\n\n"
                "For EVERY data row, extract a JSON asset record with these fields "
                "(use null when not present in the spreadsheet):\n"
                "  asset_id, serial_number, asset_type, description, location, address,\n"
                "  feeder, voltage_level, capacity_rating, manufacturer, model,\n"
                "  condition_rating, status, installation_date, last_inspection_date\n\n"
                "Map spreadsheet columns to these fields as best you can. "
                "Do NOT skip rows — include one record per data row even if some fields are null.\n\n"
                "Return ONLY a JSON object: "
                "{\"asset_records\": [{...}, ...], \"total_assets\": <int>}"
            )

            try:
                result  = _llm_call(prompt)
                records = result.get("asset_records") or []
                for rec in records:
                    rec["source_sheet"]    = sheet["sheet_name"]
                    rec["source_document"] = fe.filename
                sheet_records.extend(records)
            except Exception as e:
                logger.error(f"Asset extraction failed for sheet '{sheet['sheet_name']}' batch {batch_idx}: {e}")

        all_asset_records.extend(sheet_records)
        logger.info(f"Sheet '{sheet['sheet_name']}': {len(sheet_records)} asset records extracted "
                    f"({len(batches)} batch(es)).")

    return {
        "filename":          fe.filename,
        "filepath":          filepath,
        "document_type":     fe.document_type,
        "document_category": fe.document_category,
        "document_quality":  "spreadsheet",
        "sheets_processed":  sheets_processed,
        "sheets_skipped":    sheets_skipped if sheets_skipped else [],
        "step_extraction": {
            "asset_records":    all_asset_records,
            "total_assets":     len(all_asset_records),
            "sheets_processed": sheets_processed,
        },
    }


# ── Plain-text extraction ─────────────────────────────────────────────────────

def _extract_text_file(filepath: str) -> list:
    try:
        with open(filepath, "r", errors="replace") as f:
            content = f.read()
        return [{
            "section_number": None,
            "section_name":   "Full Document",
            "page_number":    1,
            "text":           content[:10000],
            "tables":         [],
            "images":         [],
        }]
    except Exception as e:
        return [{
            "section_number": None,
            "section_name":   "Error",
            "page_number":    0,
            "text":           str(e),
            "tables":         [],
            "images":         [],
        }]


# ── Section relevance tagging (chunked batches) ───────────────────────────────

def _tag_all_sections_chunked(sections: list, step: ProcessStepDef) -> list:
    """
    Tag sections in batches of TAGGING_BATCH_SIZE to avoid context limit failures.
    Empty and low-quality sections are skipped and given score 0.0.
    """
    tagged = list(sections)

    # Classify each section as taggable or skip.
    # Sections with tables are always taggable regardless of alpha ratio —
    # technical drawings have low alpha (codes, numbers) but rich table content.
    taggable_indices = []
    for idx, s in enumerate(sections):
        text = (s.get("text") or "").strip()
        has_tables = bool(s.get("tables"))
        if (text and _alpha_ratio(text) >= MIN_ALPHA_RATIO) or has_tables:
            taggable_indices.append(idx)
        else:
            tagged[idx] = {**sections[idx],
                           "relevance_score": 0.0,
                           "relevance_reason": "Low text quality or empty",
                           "content_tags": []}

    if not taggable_indices:
        return tagged

    # Process in batches
    for batch_start in range(0, len(taggable_indices), TAGGING_BATCH_SIZE):
        batch_indices = taggable_indices[batch_start:batch_start + TAGGING_BATCH_SIZE]
        batch_tags = _tag_batch(sections, batch_indices, step)
        for rank, idx in enumerate(batch_indices):
            tags = batch_tags[rank] if rank < len(batch_tags) else {}
            tagged[idx] = {**sections[idx],
                           "relevance_score":  tags.get("relevance_score", 0.0),
                           "relevance_reason": tags.get("relevance_reason", "Tagging failed"),
                           "content_tags":     tags.get("content_tags", [])}

    return tagged


def _tag_batch(sections: list, indices: list, step: ProcessStepDef) -> list:
    """Tag a single batch of sections with one LLM call."""
    section_blocks = []
    for rank, idx in enumerate(indices):
        s = sections[idx]
        # Use up to 600 chars of text preview
        preview = (s.get("text") or "")[:600].strip()
        # Append table headers so the LLM can see structured content in drawings
        # (technical drawings have low text but rich table data)
        table_preview = ""
        for tbl in (s.get("tables") or [])[:4]:
            headers = " | ".join(str(h) for h in (tbl.get("headers") or []) if h)
            if headers:
                table_preview += f"\n[Table: {headers}]"
        section_blocks.append(
            f"[{rank}] Section: {s.get('section_name')} (page {s.get('page_number')})\n"
            f"{preview}{table_preview}"
        )

    prompt = (
        f"You are tagging document sections for relevance to a process step.\n\n"
        f"Process step: {step.step_name}\n"
        f"Step details: {(step.details or step.summary or '').strip()[:500]}\n\n"
        f"Tag each of the {len(indices)} sections below.\n\n"
        + "\n\n---\n\n".join(section_blocks)
        + "\n\n"
        "Respond with a JSON object: {\"sections\": [<one entry per section in order>]}\n"
        "Each entry: {\"relevance_score\": 0.0-1.0, \"relevance_reason\": \"one sentence\", "
        "\"content_tags\": [\"short\", \"identifiers\"]}\n"
        "Tags: short lowercase identifiers for topics that genuinely appear in the section "
        "(e.g. 'hv', 'lv', 'sl', 'earthing', 'easement', 'substation', "
        "'funding', 'scope', 'requirements', 'schedule', 'cost', "
        "'transformer', 'voltage', 'cubicle', 'switchgear', 'siteplan'). "
        "Only include tags where the content actually contains that topic.\n"
        "Important tagging rules:\n"
        "- Tag 'funding' for any section mentioning: determination of funding, capital contribution, "
        "contestable works, non-contestable works, ancillary network services fees, reimbursement, "
        "administration fee, design fee, connection offer fee — even if the section is short.\n"
        "- Tag 'easement' for any section mentioning: land interests, easement, LIG, "
        "land interest guidelines — even if the value is 'NA' or 'nil'.\n"
        "- Tag 'siteplan' for any section containing site plan drawings, substation notes tables, "
        "electrical network diagrams, or asset annotations on a drawing.\n"
        "- Tag 'transformer' for any section mentioning transformer size, kVA, MVA ratings.\n"
        "- Tag 'switchgear' for any section mentioning HV switchgear, LV switchgear, RMU, cubicle, "
        "Siemens RLR, or switchboard specifications.\n"
        "- Tag 'voltage' for any section mentioning operating voltage level (kV).\n"
        "- Tag 'cubicle' for any section mentioning cubicle size, dimensions, or enclosure specs.\n"
        "- Assign relevance_score >= 0.5 to any section whose heading or content matches "
        "one of the above keywords, regardless of section length."
    )

    expected = len(indices)
    for attempt in range(LLM_MAX_RETRIES):
        try:
            result = _llm_call_fast(prompt)
            sections_out = result.get("sections", [])
            if len(sections_out) == expected:
                return sections_out
            # Count mismatch — pad or trim and log
            logger.warning(
                f"_tag_batch: expected {expected} section tags, got {len(sections_out)} "
                f"(attempt {attempt + 1}/{LLM_MAX_RETRIES})"
            )
            if attempt < LLM_MAX_RETRIES - 1:
                time.sleep(LLM_RETRY_DELAYS[attempt])
                continue
            # Last attempt: return what we have (pad with empty dicts if short)
            while len(sections_out) < expected:
                sections_out.append({})
            return sections_out[:expected]
        except Exception as e:
            logger.error(f"_tag_batch failed (attempt {attempt + 1}/{LLM_MAX_RETRIES}): {e}")
            if attempt < LLM_MAX_RETRIES - 1:
                time.sleep(LLM_RETRY_DELAYS[attempt])
    return [{} for _ in indices]


# ── Structured step extraction ────────────────────────────────────────────────

def _structured_extraction(relevant_sections: list, step: ProcessStepDef) -> dict:
    """
    Given the relevant sections, ask the LLM to produce a structured extraction
    matching the expected output schema of the process step.

    Uses a two-pass approach for large documents:
      Pass 1 — extract from the highest-scored sections (up to ~8000 chars)
      Pass 2 — merge any additional lower-scored sections if context allows
    """
    if not relevant_sections:
        return {"note": "No relevant sections found for this process step."}

    # Sort by relevance score descending so highest-value content goes first
    sorted_sections = sorted(relevant_sections, key=lambda s: s.get("relevance_score", 0), reverse=True)

    # Build section blocks — use more text per section (2000 chars) for better fidelity
    parts = []
    total_chars = 0
    CHAR_BUDGET = 12000  # allow more context for the extraction prompt

    for s in sorted_sections:
        tags  = s.get("content_tags", [])
        text  = (s.get("text") or "")[:2000]
        tables_detail = ""
        if s.get("tables"):
            for tbl in s["tables"][:3]:  # include up to 3 tables per section
                headers = " | ".join(tbl.get("headers", []))
                rows    = "\n".join(" | ".join(r) for r in tbl.get("rows", [])[:10])
                tables_detail += f"\nTable:\n{headers}\n{rows}"
        block = (
            f"[Section: {s.get('section_name')} | Page: {s.get('page_number')} | "
            f"Score: {s.get('relevance_score', 0):.2f} | Tags: {tags}]\n"
            f"{text}{tables_detail}"
        )
        if total_chars + len(block) > CHAR_BUDGET:
            break
        parts.append(block)
        total_chars += len(block)

    sections_block = "\n\n---\n\n".join(parts)

    # Build a focused expected output description
    expected_desc = _build_expected_desc(step)

    # For sub-step extractions, add strict cross-category exclusion instructions
    # to prevent items from one voltage class or infrastructure type bleeding into another.
    scope_instruction = ""
    if step.step_id and step.step_id.startswith("sub-step-extract-"):
        tag_hint = _derive_tag_hint(step.step_id)
        if tag_hint:
            all_categories = ["hv", "lv", "sl", "earthing", "easement", "substation", "funding"]
            other_cats = [t.upper() for t in all_categories if t != tag_hint]
            scope_instruction = (
                f"- SCOPE: Extract ONLY items that explicitly relate to {tag_hint.upper()}. "
                f"Do NOT include items belonging to other categories "
                f"({', '.join(other_cats)}) even if they appear in the same section.\n"
                f"- If an item references a different voltage class or infrastructure type "
                f"(e.g. an HV cable in an LV section), exclude it.\n"
                f"- General design notes or 'method of supply' summaries are NOT {tag_hint.upper()} "
                f"requirements unless they explicitly describe {tag_hint.upper()} infrastructure.\n"
            )

    # Build instructional guidance from instructional_sub_steps when present
    inst_guidance = ""
    if step.instructional_sub_steps:
        inst_parts = []
        for k, ist in enumerate(step.instructional_sub_steps, 1):
            title = ist.get("title", "")
            instructions = (ist.get("instructions") or "").strip()
            if title or instructions:
                inst_parts.append(f"  {k}. {title}:\n     {instructions}")
        if inst_parts:
            inst_guidance = (
                "Detailed extraction instructions (follow these carefully):\n"
                + "\n".join(inst_parts)
                + "\n\n"
            )

    prompt = (
        f"You are extracting structured engineering data from document sections.\n\n"
        f"Task: {step.step_name}\n"
        f"Context: {(step.details or step.summary or '').strip()[:800]}\n\n"
        + inst_guidance
        + (f"Required output structure:\n{expected_desc}\n\n" if expected_desc else "")
        + f"Document sections (highest relevance first):\n\n{sections_block}\n\n"
        "Instructions:\n"
        "- Extract ALL items that match the required output categories\n"
        + scope_instruction
        + "- For each item include: 'description' (verbatim or close paraphrase from source), "
        "'source_section' (section name), 'source_page' (page number)\n"
        "- Include specific values: voltages (kV), quantities, distances (m), costs ($), "
        "reference numbers, asset IDs\n"
        "- Use the exact category keys from the required output structure\n"
        "- Do NOT invent data not present in the source text\n"
        "- Do NOT include metadata from prior workflow steps (no 'documents' list, no 'processing_plan')\n"
        "Return a JSON object with only the extraction categories."
    )

    try:
        return _llm_call(prompt)
    except Exception as e:
        logger.error(f"_structured_extraction failed after all retries: {e}")
        return {"error": str(e)}


def _build_expected_desc(step: ProcessStepDef) -> str:
    """Build a concise expected output description from the step definition."""
    if not step.expected_output:
        return ""
    fields = step.expected_output.get("fields")
    description = step.expected_output.get("description", "")
    # Only use fields dict if it looks like extraction categories (not pipeline schemas)
    pipeline_keys = {"documents", "scan_results", "processing_plan", "total_documents",
                     "matched_files", "total_files_scanned", "output_file"}
    if isinstance(fields, dict):
        clean_fields = {k: v for k, v in fields.items() if k not in pipeline_keys}
        if clean_fields:
            return json.dumps(clean_fields, indent=2)
    if description:
        # Strip pipeline-step references from the description
        lines = [l for l in description.split("|")
                 if not any(kw in l.lower() for kw in ["step 1", "step 2", "documentreview", "processingplan"])]
        return " | ".join(lines).strip()
    return ""


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8090)
